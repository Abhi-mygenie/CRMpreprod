"""
Iteration 19 tests for CR-063 (Customer Detail whatsapp opt-in badge + edit modal switch)
and CR-065 (Message Status resend indicator + export columns).

Scope:
- CR-063 backend contract: GET /customers/{id} returns whatsapp_opt_in field;
  PUT /customers/{id} persists whatsapp_opt_in true/false without wiping other fields.
- CR-065 export headers: CSV & XLSX final columns are "Resend Count" and "Last Resend At".
- CR-065 seed row surfaces in /message-logs list with resend_count=2.

Cleanup:
- Any seeded whatsapp_message_logs row is removed at teardown.
- Toggled customer opt-in is restored to original value.
"""

import os
import io
import uuid
from datetime import datetime, timezone, timedelta

import pytest
import requests
from openpyxl import load_workbook
from pymongo import MongoClient

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://full-stack-crm-1.preview.emergentagent.com").rstrip("/")
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mygenie_admin:QplazmMzalpq@52.66.232.149:27017/mygenie")
DB_NAME = os.environ.get("DB_NAME", "mygenie")
OWNER_EMAIL = "owner@jehsnest.com"
OWNER_PASSWORD = "Qplazm@10"


# ---------- fixtures ----------
@pytest.fixture(scope="module")
def token():
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": OWNER_EMAIL, "password": OWNER_PASSWORD},
        timeout=30,
    )
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text[:200]}"
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def db():
    client = MongoClient(MONGO_URL)
    return client[DB_NAME]


@pytest.fixture(scope="module")
def owner_user_id(db):
    u = db.users.find_one({"email": OWNER_EMAIL}, {"id": 1})
    assert u, "owner user not found"
    return u["id"]


@pytest.fixture(scope="module")
def sample_customer_id(db, owner_user_id):
    """Pick an existing customer for the CR-063 toggle test."""
    c = db.customers.find_one({"user_id": owner_user_id}, {"id": 1, "whatsapp_opt_in": 1, "name": 1, "phone": 1})
    assert c, "no customer available for owner"
    return c["id"]


# ---------- CR-063 · backend contract ----------
class TestCR063CustomerOptIn:
    def test_get_customer_includes_opt_in(self, auth_headers, sample_customer_id):
        r = requests.get(f"{BASE_URL}/api/customers/{sample_customer_id}", headers=auth_headers, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "whatsapp_opt_in" in data
        assert isinstance(data["whatsapp_opt_in"], bool)

    def test_toggle_opt_in_persists_and_restores(self, auth_headers, sample_customer_id):
        # snapshot
        r0 = requests.get(f"{BASE_URL}/api/customers/{sample_customer_id}", headers=auth_headers, timeout=30)
        assert r0.status_code == 200
        orig = r0.json()
        original_opt = orig.get("whatsapp_opt_in", True)
        original_name = orig.get("name")
        original_phone = orig.get("phone")

        new_val = not original_opt
        try:
            # toggle
            r1 = requests.put(
                f"{BASE_URL}/api/customers/{sample_customer_id}",
                headers=auth_headers,
                json={"whatsapp_opt_in": new_val},
                timeout=30,
            )
            assert r1.status_code == 200, r1.text[:300]

            # verify persisted
            r2 = requests.get(f"{BASE_URL}/api/customers/{sample_customer_id}", headers=auth_headers, timeout=30)
            assert r2.status_code == 200
            after = r2.json()
            assert after["whatsapp_opt_in"] == new_val, f"expected {new_val}, got {after['whatsapp_opt_in']}"
            # Other fields untouched
            assert after.get("name") == original_name, "name got mutated"
            assert after.get("phone") == original_phone, "phone got mutated"
        finally:
            # restore
            requests.put(
                f"{BASE_URL}/api/customers/{sample_customer_id}",
                headers=auth_headers,
                json={"whatsapp_opt_in": original_opt},
                timeout=30,
            )
            r3 = requests.get(f"{BASE_URL}/api/customers/{sample_customer_id}", headers=auth_headers, timeout=30)
            assert r3.json().get("whatsapp_opt_in") == original_opt, "restore failed"


# ---------- CR-065 · seed a resent log and verify it surfaces via /message-logs ----------
class TestCR065SeededResend:
    SEEDED_ID = None

    def test_seed_and_list(self, auth_headers, db, owner_user_id):
        now = datetime.now(timezone.utc)
        doc = {
            "id": str(uuid.uuid4()),
            "user_id": owner_user_id,
            "customer_name": "QA Resent Test",
            "customer_phone": "9000003001",
            "template_name": "qa_test",
            "status": "pending",
            "is_test": False,
            "created_at": (now - timedelta(hours=2)).isoformat(),
            "resend_count": 2,
            "last_resend_at": now.isoformat(),
            "status_history": [],
        }
        db.whatsapp_message_logs.insert_one(doc)
        TestCR065SeededResend.SEEDED_ID = doc["id"]
        try:
            # search by phone to isolate this row in the paginated list
            r = requests.get(
                f"{BASE_URL}/api/whatsapp/message-logs?search=9000003001&limit=10",
                headers=auth_headers,
                timeout=30,
            )
            assert r.status_code == 200, r.text[:300]
            payload = r.json()
            # response shape: dict with 'logs' or plain list
            items = payload.get("logs") if isinstance(payload, dict) else payload
            assert items, f"seeded log not returned: {payload}"
            match = next((l for l in items if l.get("id") == doc["id"]), None)
            assert match is not None, f"seeded id {doc['id']} not in returned logs"
            assert match.get("resend_count") == 2
            assert match.get("last_resend_at"), "last_resend_at missing"
        finally:
            db.whatsapp_message_logs.delete_one({"id": TestCR065SeededResend.SEEDED_ID})

    def test_seed_row_removed(self, db):
        assert TestCR065SeededResend.SEEDED_ID is not None
        left = db.whatsapp_message_logs.count_documents({"id": TestCR065SeededResend.SEEDED_ID})
        assert left == 0, "seed cleanup failed"


# ---------- CR-065 · export column additions ----------
class TestCR065ExportHeaders:
    def test_csv_export_headers_end_with_resend_columns(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/whatsapp/message-logs/export?format=csv",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        first_line = r.text.splitlines()[0]
        cols = first_line.split(",")
        assert cols[-2] == "Resend Count", f"cols={cols}"
        assert cols[-1] == "Last Resend At", f"cols={cols}"

    def test_xlsx_export_returns_200_and_has_new_headers(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/whatsapp/message-logs/export?format=xlsx",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "").lower()
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [c.value for c in ws[1]]
        assert headers_row[-2] == "Resend Count"
        assert headers_row[-1] == "Last Resend At"


# ---------- Regression: filters still respond ----------
class TestRegressionMessageLogs:
    def test_message_logs_list_ok(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/whatsapp/message-logs?limit=5", headers=auth_headers, timeout=30)
        assert r.status_code == 200

    def test_message_logs_with_status_filter(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/whatsapp/message-logs?status=delivered&limit=5", headers=auth_headers, timeout=30)
        assert r.status_code == 200
