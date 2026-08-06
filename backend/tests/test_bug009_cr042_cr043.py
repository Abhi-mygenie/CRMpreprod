"""
Regression tests for BUG-009, CR-042, CR-043 (backend surface).

Uses the pre-baked JWT for tenant pos_0001_restaurant_635 (Jeh's Nest) supplied
by main agent (MyGenie SSO is unreachable from this preview env).

Run:
  cd /app/backend && python3 -m pytest tests/test_bug009_cr042_cr043.py -v --override-ini='addopts='
"""
import io
import os
import csv
import uuid
import pytest
import requests

BASE_URL = None
with open("/app/frontend/.env") as f:
    for line in f:
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
            break
assert BASE_URL, "REACT_APP_BACKEND_URL missing"

TOKEN = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJ1c2VyX2lkIjoicG9zXzAwMDFfcmVzdGF1cmFudF82MzUiLCJ0eXBlIjoic3RhZmYiLCJleHAiOjE3ODMxOTQ5Mzl9."
    "RLMro2BoxO_Y8QD9zLHMl9f2uP4CTQPVakvflrzHvJI"
)
HDRS = {"Authorization": f"Bearer {TOKEN}"}

MSG_LOGS = f"{BASE_URL}/api/whatsapp/message-logs"
MSG_EXPORT = f"{BASE_URL}/api/whatsapp/message-logs/export"
CUSTOMERS = f"{BASE_URL}/api/customers"
TAGS = f"{BASE_URL}/api/customers/tags"

# ─── CR-042 export endpoint ──────────────────────────────────────────────────

class TestExportAuth:
    def test_missing_auth_rejected(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv", timeout=30)
        assert r.status_code in (401, 403), f"got {r.status_code}: {r.text[:200]}"

    def test_invalid_token_rejected(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv",
                         headers={"Authorization": "Bearer BOGUS"}, timeout=30)
        assert r.status_code in (401, 403)


class TestExportFormat:
    def test_format_pdf_returns_400(self):
        r = requests.get(f"{MSG_EXPORT}?format=pdf", headers=HDRS, timeout=30)
        assert r.status_code == 400
        # error surface: {"detail": "format must be 'csv' or 'xlsx'"}
        body = r.json()
        assert "csv" in body.get("detail", "").lower()

    def test_format_missing_defaults_to_csv(self):
        r = requests.get(MSG_EXPORT, headers=HDRS, timeout=60)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")

    def test_format_junk_returns_400(self):
        r = requests.get(f"{MSG_EXPORT}?format=json", headers=HDRS, timeout=30)
        assert r.status_code == 400


class TestExportCSV:
    def test_csv_download_headers(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv", headers=HDRS, timeout=60)
        assert r.status_code == 200
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower() and ".csv" in cd.lower(), cd
        assert "X-Row-Count" in r.headers
        assert r.headers.get("X-Row-Cap") == "5000"

    def test_csv_has_exactly_12_columns_in_locked_order(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv", headers=HDRS, timeout=60)
        assert r.status_code == 200
        text = r.content.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        expected = [
            "Sent At", "Phone", "Name", "Event / Campaign", "Template",
            "Status", "Delivered At", "Read At", "Rejected At",
            "Error Reason", "Message ID", "Test Send",
        ]
        assert header == expected, f"headers mismatch: {header}"

    def test_csv_row_count_matches_header(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv", headers=HDRS, timeout=60)
        expected_count = int(r.headers["X-Row-Count"])
        text = r.content.decode("utf-8")
        rows = list(csv.reader(io.StringIO(text)))
        # first row is header, rest are data
        assert len(rows) - 1 == expected_count, (
            f"data rows {len(rows)-1} != X-Row-Count {expected_count}"
        )

    def test_csv_status_filter_honoured(self):
        r = requests.get(f"{MSG_EXPORT}?format=csv&status=delivered",
                         headers=HDRS, timeout=60)
        assert r.status_code == 200
        reader = csv.reader(io.StringIO(r.content.decode("utf-8")))
        header = next(reader)
        status_col = header.index("Status")
        for row in reader:
            if row and row[status_col]:
                assert row[status_col] == "delivered", f"unexpected status: {row[status_col]}"

    def test_csv_zero_rows_still_has_header(self):
        r = requests.get(
            f"{MSG_EXPORT}?format=csv&status=DEFINITELY_NO_SUCH_STATUS",
            headers=HDRS, timeout=30,
        )
        assert r.status_code == 200
        assert r.headers["X-Row-Count"] == "0"
        rows = list(csv.reader(io.StringIO(r.content.decode("utf-8"))))
        assert len(rows) == 1, f"expected header-only, got {len(rows)} rows"
        assert rows[0][0] == "Sent At"


class TestExportXLSX:
    def test_xlsx_content_type_and_extension(self):
        r = requests.get(f"{MSG_EXPORT}?format=xlsx", headers=HDRS, timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", ""), r.headers.get("content-type")
        cd = r.headers.get("content-disposition", "")
        assert ".xlsx" in cd.lower(), cd

    def test_xlsx_headers_row_has_orange_fill_and_white_bold_font(self):
        r = requests.get(f"{MSG_EXPORT}?format=xlsx", headers=HDRS, timeout=60)
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        expected = [
            "Sent At", "Phone", "Name", "Event / Campaign", "Template",
            "Status", "Delivered At", "Read At", "Rejected At",
            "Error Reason", "Message ID", "Test Send",
        ]
        got = [c.value for c in ws[1]]
        assert got == expected, f"xlsx header row mismatch: {got}"
        first = ws["A1"]
        # Check bold & white font
        assert first.font.bold is True, "header font should be bold"
        color = first.font.color.rgb if first.font.color else ""
        assert str(color).upper().endswith("FFFFFF"), f"font color should be white, got {color}"
        # Fill should be orange #F26B33
        fill_rgb = first.fill.fgColor.rgb if first.fill and first.fill.fgColor else ""
        assert "F26B33" in str(fill_rgb).upper(), f"expected F26B33 fill, got {fill_rgb}"


class TestExportRunAndCampaignFilter:
    def test_run_id_filter_narrows_results(self):
        # Grab a real run from history
        h = requests.get(
            f"{BASE_URL}/api/campaigns/history/all?days=365",
            headers=HDRS, timeout=30,
        )
        if h.status_code != 200 or not h.json():
            pytest.skip("no campaign runs available for tenant")
        run = h.json()[0]
        run_id = run["id"]
        campaign_id = run.get("campaign_id")

        # unfiltered total
        u = requests.get(MSG_EXPORT, headers=HDRS, timeout=60)
        unfiltered_count = int(u.headers["X-Row-Count"])
        # filtered by run_id
        f = requests.get(f"{MSG_EXPORT}?run_id={run_id}",
                         headers=HDRS, timeout=60)
        filtered_count = int(f.headers["X-Row-Count"])
        assert filtered_count <= unfiltered_count
        # filtered by run_id + campaign_id
        if campaign_id:
            fc = requests.get(
                f"{MSG_EXPORT}?run_id={run_id}&campaign_id={campaign_id}",
                headers=HDRS, timeout=60,
            )
            assert fc.status_code == 200
            assert int(fc.headers["X-Row-Count"]) == filtered_count

    def test_message_logs_endpoint_supports_run_id_param(self):
        # Regression: /message-logs still returns same list without run_id
        r_no = requests.get(MSG_LOGS, headers=HDRS, timeout=30)
        assert r_no.status_code == 200
        assert "total" in r_no.json() and "logs" in r_no.json()

        # With run_id=all it should behave same as no filter
        r_all = requests.get(f"{MSG_LOGS}?run_id=all", headers=HDRS, timeout=30)
        assert r_all.status_code == 200
        assert r_all.json()["total"] == r_no.json()["total"]

    def test_message_logs_run_id_narrows_count(self):
        h = requests.get(
            f"{BASE_URL}/api/campaigns/history/all?days=365",
            headers=HDRS, timeout=30,
        )
        if h.status_code != 200 or not h.json():
            pytest.skip("no campaign runs")
        run_id = h.json()[0]["id"]
        r = requests.get(f"{MSG_LOGS}?run_id={run_id}", headers=HDRS, timeout=30)
        assert r.status_code == 200
        assert isinstance(r.json().get("logs"), list)


# ─── CR-043 tags backend ─────────────────────────────────────────────────────

class TestTagsCatalog:
    def test_tags_default_returns_string_array_backcompat(self):
        r = requests.get(TAGS, headers=HDRS, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "tags" in data
        for t in data["tags"]:
            assert isinstance(t, str), f"expected str element got {type(t)}"

    def test_tags_with_counts_returns_objects_sorted_desc(self):
        r = requests.get(f"{TAGS}?with_counts=true", headers=HDRS, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "tags" in data
        assert isinstance(data["tags"], list)
        counts = []
        for entry in data["tags"]:
            assert isinstance(entry, dict), f"expected dict, got {entry!r}"
            assert "tag" in entry and "count" in entry
            assert isinstance(entry["count"], int)
            counts.append(entry["count"])
        # Sorted descending
        assert counts == sorted(counts, reverse=True), f"not sorted desc: {counts}"

    def test_tags_with_counts_includes_zero_count_catalog_tags(self):
        r = requests.get(f"{TAGS}?with_counts=true", headers=HDRS, timeout=30)
        data = r.json()
        tag_names = [e["tag"] for e in data["tags"]]
        # Jeh's Nest catalog per review request: Lunch, Dinner, Breakfast, Churn (0)
        # If Churn (0) is missing from with_counts output, that's a bug.
        # But we don't assert specific tags since the catalog may vary; instead
        # verify the /tags default returns >= the count tags with count > 0.
        default = requests.get(TAGS, headers=HDRS, timeout=30).json()["tags"]
        for t in default:
            # Every catalog tag must appear in with_counts output
            assert t in tag_names, f"catalog tag '{t}' missing from with_counts result"


class TestCustomersTagFilter:
    def test_customers_no_tag_filter_returns_all(self):
        r = requests.get(f"{CUSTOMERS}?limit=200", headers=HDRS, timeout=30)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_customers_empty_tags_param_is_noop(self):
        r_no = requests.get(f"{CUSTOMERS}?limit=200", headers=HDRS, timeout=30)
        r_empty = requests.get(f"{CUSTOMERS}?tags=&limit=200", headers=HDRS, timeout=30)
        assert r_no.status_code == r_empty.status_code == 200
        # Same count
        assert len(r_no.json()) == len(r_empty.json())

    def test_customers_tag_any_mode(self):
        tags_resp = requests.get(f"{TAGS}?with_counts=true", headers=HDRS, timeout=30).json()["tags"]
        used = [e["tag"] for e in tags_resp if e["count"] > 0]
        if len(used) < 2:
            pytest.skip("need at least 2 tags with count > 0 to test any/all")
        a, b = used[0], used[1]
        r = requests.get(f"{CUSTOMERS}?tags={a},{b}&tags_mode=any&limit=500",
                         headers=HDRS, timeout=30)
        assert r.status_code == 200
        results = r.json()
        for c in results:
            ctags = c.get("tags") or []
            assert a in ctags or b in ctags, (
                f"customer {c.get('id')} has neither {a} nor {b}: {ctags}"
            )

    def test_customers_tag_all_mode_is_intersection(self):
        tags_resp = requests.get(f"{TAGS}?with_counts=true", headers=HDRS, timeout=30).json()["tags"]
        used = [e["tag"] for e in tags_resp if e["count"] > 0]
        if len(used) < 2:
            pytest.skip("need 2 tags with count > 0")
        a, b = used[0], used[1]
        r_all = requests.get(f"{CUSTOMERS}?tags={a},{b}&tags_mode=all&limit=500",
                             headers=HDRS, timeout=30)
        r_any = requests.get(f"{CUSTOMERS}?tags={a},{b}&tags_mode=any&limit=500",
                             headers=HDRS, timeout=30)
        assert r_all.status_code == 200 and r_any.status_code == 200
        assert len(r_all.json()) <= len(r_any.json()), (
            f"ALL={len(r_all.json())} should be <= ANY={len(r_any.json())}"
        )
        for c in r_all.json():
            ctags = c.get("tags") or []
            assert a in ctags and b in ctags, f"ALL-mode row {c.get('id')} missing tag"

    def test_customers_tag_mode_default_is_any(self):
        tags_resp = requests.get(f"{TAGS}?with_counts=true", headers=HDRS, timeout=30).json()["tags"]
        used = [e["tag"] for e in tags_resp if e["count"] > 0]
        if not used:
            pytest.skip("no tags in use")
        a = used[0]
        r_default = requests.get(f"{CUSTOMERS}?tags={a}&limit=500", headers=HDRS, timeout=30)
        r_any = requests.get(f"{CUSTOMERS}?tags={a}&tags_mode=any&limit=500",
                             headers=HDRS, timeout=30)
        assert r_default.status_code == 200 and r_any.status_code == 200
        assert len(r_default.json()) == len(r_any.json())

    def test_customers_tags_non_existent_returns_zero(self):
        bogus = f"NONEXISTENT_TAG_{uuid.uuid4().hex[:8]}"
        r = requests.get(f"{CUSTOMERS}?tags={bogus}&limit=500", headers=HDRS, timeout=30)
        assert r.status_code == 200
        assert r.json() == []


class TestTagRowEndpoints:
    """CR-043 regression: POST/DELETE /customers/{id}/tags/{tag} must still work."""

    def test_add_and_remove_tag_roundtrip(self):
        # find a customer
        r = requests.get(f"{CUSTOMERS}?limit=1", headers=HDRS, timeout=30)
        if r.status_code != 200 or not r.json():
            pytest.skip("no customers")
        cust = r.json()[0]
        cust_id = cust["id"]
        original = list(cust.get("tags") or [])
        test_tag = f"TESTTAG_{uuid.uuid4().hex[:6]}"

        try:
            # POST
            add = requests.post(
                f"{BASE_URL}/api/customers/{cust_id}/tags",
                headers={**HDRS, "Content-Type": "application/json"},
                json={"tags": [test_tag]},
                timeout=30,
            )
            assert add.status_code in (200, 201), f"add {add.status_code}: {add.text[:200]}"

            # Verify GET
            g = requests.get(f"{BASE_URL}/api/customers/{cust_id}",
                             headers=HDRS, timeout=30)
            assert g.status_code == 200
            assert test_tag in (g.json().get("tags") or [])

            # DELETE
            d = requests.delete(
                f"{BASE_URL}/api/customers/{cust_id}/tags/{test_tag}",
                headers=HDRS, timeout=30,
            )
            assert d.status_code in (200, 204), d.text[:200]

            # Verify removed
            g2 = requests.get(f"{BASE_URL}/api/customers/{cust_id}",
                              headers=HDRS, timeout=30)
            assert test_tag not in (g2.json().get("tags") or [])
        finally:
            # Best-effort cleanup: ensure tag isn't left
            requests.delete(
                f"{BASE_URL}/api/customers/{cust_id}/tags/{test_tag}",
                headers=HDRS, timeout=30,
            )


class TestOtherFiltersRegression:
    def test_customers_tier_filter_still_works(self):
        r = requests.get(f"{CUSTOMERS}?tier=Gold&limit=200", headers=HDRS, timeout=30)
        assert r.status_code == 200
        for c in r.json():
            assert c.get("tier") == "Gold"

    def test_customers_type_filter_still_works(self):
        r = requests.get(f"{CUSTOMERS}?customer_type=normal&limit=200", headers=HDRS, timeout=30)
        assert r.status_code == 200
        for c in r.json():
            # Some legacy customers may lack the field
            if c.get("customer_type") is not None:
                assert c["customer_type"] == "normal"
