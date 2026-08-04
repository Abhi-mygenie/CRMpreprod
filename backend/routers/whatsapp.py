from fastapi import APIRouter, HTTPException, Depends, Query, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse   # CR-042: message report streaming download
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
import hashlib
import hmac
import json
import logging
import os
import re
import shutil
import time
import uuid
from io import BytesIO
from urllib.parse import parse_qs
import httpx

from core.database import db
from core.auth import get_current_user
from core.whatsapp import send_single_message, WhatsAppMessage, log_message_attempt
from core.whatsapp_status import next_status
from core.whatsapp_variables import WHATSAPP_VARIABLES
from core.meta_media import upload_to_meta_uploads
from core.s3 import put_public_object, get_public_url, S3_CONFIGURED
from models.schemas import (
    AUTOMATION_EVENTS, POS_EVENTS, CRM_EVENTS
)

logger = logging.getLogger(__name__)

# Message status constants
# CR-036 B.2 (E-B2-1): add "failed" status for G5 fail-loud rows (media_missing)
MESSAGE_STATUSES = ["pending", "delivered", "read", "rejected", "failed"]


# ── CR-061: CRM templates env-gated authoring ──────────────────────────
# EMPTY allowlist = feature disabled for ALL tenants (safe default). Owner adds
# specific restaurant_ids to CRM_TEMPLATES_ALLOWED_RESTAURANT_IDS at deploy.
# Gate covers AUTHORING endpoints only (POST/PUT/DELETE/PATCH/submit/create/upload);
# GET custom-templates + status stay open (read-only, tenant-scoped).
# Silent 403 by owner decision (q4-lock option b) — the frontend hides the
# entry points so this API is not called by allowed tenants' UI.

def _crm_templates_allowlist() -> set:
    raw = os.environ.get("CRM_TEMPLATES_ALLOWED_RESTAURANT_IDS", "") or ""
    return {p.strip() for p in raw.split(",") if p.strip()}


def _crm_templates_enabled(user: dict) -> bool:
    allow = _crm_templates_allowlist()
    if not allow:
        return False
    rid = user.get("restaurant_id")
    if rid is None:
        return False
    return str(rid) in allow


def _require_crm_templates_enabled(user: dict) -> None:
    """CR-061: raise silent 403 when the tenant is not in the allowlist.
    Frontend hides the UI, so this should never fire for legitimate users."""
    if not _crm_templates_enabled(user):
        raise HTTPException(status_code=403, detail="Forbidden")


class TestTemplateRequest(BaseModel):
    template_id: str
    phone: str
    country_code: str = "91"
    body_values: Dict[str, str] = {}
    button_values: Optional[Dict[str, str]] = None  # CR-069: button URL suffix values
    media_url: Optional[str] = None
    media_filename: Optional[str] = None

router = APIRouter(prefix="/whatsapp", tags=["WhatsApp"])


# CR-004 P1: Canonical variables endpoint (replaces hardcoded frontend lists)
@router.get("/variables")
async def list_template_variables():
    """Return canonical template variables list (Phase 1: 10 vars, flat)."""
    return {"variables": WHATSAPP_VARIABLES}


@router.get("/automation/events")
async def get_automation_events():
    """Get all available automation event types with descriptions, categorized by POS and CRM"""
    
    # POS Events descriptions
    pos_event_descriptions = {
        "new_order_customer": "Notify customer when a new order is placed",
        "new_order_outlet": "Alert outlet/restaurant when a new order is received",
        "order_confirmed": "Confirm order to customer when outlet accepts",
        "order_ready_customer": "Notify customer when order is ready for pickup/serve",
        "item_ready": "Notify customer when a specific item is ready",
        "order_served": "Notify customer when order has been served",
        "item_served": "Notify customer when a specific item has been served",
        "order_ready_delivery": "Alert delivery boy when order is ready for pickup",
        "order_dispatched": "Notify customer when order is out for delivery",
        "send_bill_manual": "Manually send bill/receipt to customer",
        "send_bill_auto": "Automatically send bill after order completion",
    }
    
    # CRM Events descriptions
    crm_event_descriptions = {
        "reset_password": "Send OTP for forgot password verification",
        "welcome_message": "Welcome message for new customers",
        "birthday": "Send birthday wishes to customers",
        "anniversary": "Send anniversary wishes to customers",
        "points_earned": "Notify when customer earns loyalty points",
        "points_expiring": "Remind customers before their points expire",
        "feedback_request": "Request feedback from customers after visit",
        "send_bill": "Send bill/receipt to customer after order",
        "tier_upgrade": "Congratulate customer on loyalty tier upgrade",
        "coupon_earned": "Notify customer when they earn a coupon",
        "wallet_credit": "Confirm wallet top-up to customer",
        "wallet_debit": "Confirm wallet payment to customer",
        "bonus_points": "Notify customer of bonus points awarded",
        "points_redeemed": "Confirm points redemption to customer",
        "coupon_expiring": "Remind customers about expiring coupons",
        "inactive_customer": "Win-back message for inactive customers",
    }
    
    # Combined for backward compatibility
    event_descriptions = {**pos_event_descriptions, **crm_event_descriptions}
    
    return {
        "events": AUTOMATION_EVENTS,
        "descriptions": event_descriptions,
        "pos_events": POS_EVENTS,
        "crm_events": CRM_EVENTS,
        "pos_descriptions": pos_event_descriptions,
        "crm_descriptions": crm_event_descriptions
    }

@router.get("/api-key")
async def get_whatsapp_api_key(user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one(
        {"id": user["id"]}, 
        {"_id": 0, "authkey_api_key": 1, "brand_number": 1, "meta_waba_id": 1, "meta_access_token": 1, "meta_app_id": 1}
    )
    # CR-061: expose gating flag so frontend can hide CRM template authoring UI
    crm_templates_enabled = _crm_templates_enabled(user)
    if not user_doc:
        return {
            "authkey_api_key": "", "brand_number": "", "meta_waba_id": "",
            "meta_access_token": "", "meta_app_id": "",
            "crm_templates_enabled": crm_templates_enabled,   # CR-061
        }
    return {
        "authkey_api_key": user_doc.get("authkey_api_key", ""),
        "brand_number": user_doc.get("brand_number", ""),
        "meta_waba_id": user_doc.get("meta_waba_id", ""),
        "meta_access_token": user_doc.get("meta_access_token", ""),
        "meta_app_id": user_doc.get("meta_app_id", ""),  # CR-036 Q14 (per-tenant)
        "crm_templates_enabled": crm_templates_enabled,   # CR-061
    }

@router.put("/api-key")
async def save_whatsapp_api_key(payload: dict, user: dict = Depends(get_current_user)):
    update_fields = {}
    if "authkey_api_key" in payload:
        update_fields["authkey_api_key"] = payload.get("authkey_api_key", "")
    if "brand_number" in payload:
        update_fields["brand_number"] = payload.get("brand_number", "")
    if "meta_waba_id" in payload:
        update_fields["meta_waba_id"] = payload.get("meta_waba_id", "")
    if "meta_access_token" in payload:
        update_fields["meta_access_token"] = payload.get("meta_access_token", "")
    # CR-036 Q14 · per-tenant Meta APP_ID for /uploads endpoint (Batch B.1)
    if "meta_app_id" in payload:
        update_fields["meta_app_id"] = payload.get("meta_app_id", "").strip()
    
    if update_fields:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": update_fields}
        )
    return {"message": "WhatsApp settings saved", **update_fields}


@router.get("/authkey-templates")
async def get_authkey_templates(user: dict = Depends(get_current_user)):
    """Fetch WhatsApp templates from AuthKey.io using the user's saved API key."""
    user_doc = await db.users.find_one({"id": user["id"]}, {"_id": 0, "authkey_api_key": 1})
    api_key = user_doc.get("authkey_api_key", "") if user_doc else ""
    if not api_key:
        raise HTTPException(status_code=400, detail="WhatsApp API key not configured. Please add it in Settings.")
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            os.environ['AUTHKEY_TEMPLATES_URL'],
            headers={"Authorization": f"Basic {api_key}", "Content-Type": "application/json"},
            json={"channel": "whatsapp"},
        )
    data = resp.json()
    if not data.get("status"):
        raise HTTPException(status_code=400, detail="Invalid API key or AuthKey.io request failed.")
    templates = data.get("data", []) or []
    # CR-036 B.2 (E-B2-4): enrich AuthKey templates with media-header metadata
    # from custom_templates so the campaign wizard can render the media-missing
    # soft block (GAP-B2-1). Additive keys only — existing shape untouched.
    customs = await db.custom_templates.find(
        {"user_id": user["id"]},
        {"_id": 0, "authkey_wid": 1, "header_type": 1, "send_media_url": 1, "needs_media_reupload": 1, "buttons": 1},
    ).to_list(200)
    by_wid = {str(c["authkey_wid"]): c for c in customs if c.get("authkey_wid")}
    for t in templates:
        c = by_wid.get(str(t.get("wid", "")))
        if c:
            t["header_type"] = c.get("header_type")
            t["has_send_media"] = bool(c.get("send_media_url"))
            t["needs_media_reupload"] = bool(c.get("needs_media_reupload"))
            # CR-069: enrich with button data from custom_templates
            if c.get("buttons"):
                t["buttons"] = c["buttons"]
    return {"templates": templates}


# ---- Custom Template CRUD ----

@router.post("/custom-templates")
async def create_custom_template(payload: dict, user: dict = Depends(get_current_user)):
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    """Create a new custom WhatsApp template (saved locally as Draft)."""
    now = datetime.now(timezone.utc).isoformat()
    template_id = str(uuid.uuid4())
    
    # Extract variables from body
    import re
    body = payload.get("body", "")
    variables = list(set(re.findall(r'\{\{\d+\}\}', body)))
    variables.sort(key=lambda v: int(v.strip('{}') or 0))
    
    doc = {
        "id": template_id,
        "user_id": user["id"],
        "template_name": payload.get("template_name", "").strip(),
        "category": payload.get("category", "utility"),
        "language": payload.get("language", "en"),
        "header_type": payload.get("header_type", "none"),
        "header_content": payload.get("header_content", ""),
        "body": body,
        "footer": payload.get("footer", ""),
        "buttons": payload.get("buttons", []),
        "media_url": payload.get("media_url", ""),
        # CR-036 B.1: Meta opaque handle for approval submission
        "header_handle": payload.get("header_handle") or None,
        # CR-036 B.1: public S3 URL for send-time delivery
        "send_media_url": payload.get("send_media_url") or None,
        "send_media_filename": payload.get("send_media_filename") or None,
        "header_media_mime": payload.get("header_media_mime") or None,
        "needs_media_reupload": False,
        "variables": variables,
        "status": "draft",
        "created_at": now,
        "updated_at": now
    }
    
    await db.custom_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc


# CR-036 Batch B.1: Dual upload (Meta /uploads + S3) for media header files
_MEDIA_CAPS = {
    "image":    {"max_bytes": 5 * 1024 * 1024,   "mimes": {"image/jpeg", "image/png"}},
    "video":    {"max_bytes": 16 * 1024 * 1024,   "mimes": {"video/mp4", "video/3gpp"}},
    "document": {"max_bytes": 100 * 1024 * 1024,  "mimes": {"application/pdf"}},
}

def _classify_mime(mime: str) -> Optional[str]:
    for kind, cfg in _MEDIA_CAPS.items():
        if mime in cfg["mimes"]:
            return kind
    return None


# CR-036 B.3 (E-B3-1): shared Meta+S3 processing, used by single-shot and chunked flows
async def _process_media_upload(user: dict, contents: bytes, mime: str, filename: str, template_slug: str) -> dict:
    # 1. Fetch user's Meta creds
    user_doc = await db.users.find_one(
        {"id": user["id"]},
        {"meta_waba_id": 1, "meta_access_token": 1, "meta_app_id": 1},
    )
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    waba_id = (user_doc.get("meta_waba_id") or "").strip()
    access_token = (user_doc.get("meta_access_token") or "").strip()
    if not waba_id or not access_token:
        raise HTTPException(
            status_code=400,
            detail="Meta credentials missing — configure Settings > WhatsApp > Meta API.",
        )

    # 2. Validate MIME + size
    kind = _classify_mime(mime)
    if not kind:
        raise HTTPException(status_code=400, detail=f"Unsupported media type: {mime}")
    cap = _MEDIA_CAPS[kind]["max_bytes"]
    if len(contents) > cap:
        max_mb = cap // (1024 * 1024)
        raise HTTPException(status_code=413, detail=f"File too large. Max for {kind}: {max_mb} MB.")

    # 3. Upload to Meta /uploads (Part 1 — approval handle)
    user_for_meta = {**user_doc, "meta_access_token": access_token}
    handle = await upload_to_meta_uploads(user_for_meta, contents, mime, filename or "media")

    # 4. Upload to S3 (Part 2 — delivery URL)
    if not S3_CONFIGURED:
        raise HTTPException(status_code=503, detail="S3 not configured — cannot store delivery media.")
    def _slug(s, mx=40):
        return re.sub(r"[^A-Za-z0-9._-]+", "_", (s or ""))[:mx]
    ts = int(time.time())
    s3_key = f"media-headers/{_slug(user['id'])}/{_slug(template_slug)}/{ts}_{_slug(filename or 'media')}"
    put_public_object(s3_key, contents, mime)
    send_media_url = get_public_url(s3_key)

    return {
        "handle": handle,
        "send_media_url": send_media_url,
        "mime": mime,
        "filename": filename or "media",
        "kind": kind,
    }


@router.post("/upload-media-header")
async def upload_media_header(
    file: UploadFile = File(...),
    template_slug: str = Form("template"),
    user: dict = Depends(get_current_user),
):
    """CR-036 B.1: Upload a media file to both Meta /uploads (for approval handle)
    and S3 (for send-time delivery URL). Returns both artifacts."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    contents = await file.read()
    return await _process_media_upload(user, contents, file.content_type or "", file.filename or "media", template_slug)


# CR-036 B.3 (E-B3-2…4): chunked upload flow for files > 4 MB
_MEDIA_STAGING_ROOT = "/tmp/media_uploads"
_MEDIA_CHUNK_SIZE = 4 * 1024 * 1024
_MEDIA_STAGING_TTL_SECONDS = 2 * 60 * 60


def _staging_dir(user_id: str, upload_id: str) -> str:
    safe = lambda s: re.sub(r"[^A-Za-z0-9._-]+", "_", s or "")
    return os.path.join(_MEDIA_STAGING_ROOT, safe(user_id), safe(upload_id))


def _sweep_stale_staging():
    if not os.path.isdir(_MEDIA_STAGING_ROOT):
        return
    cutoff = time.time() - _MEDIA_STAGING_TTL_SECONDS
    for user_dir in os.listdir(_MEDIA_STAGING_ROOT):
        upath = os.path.join(_MEDIA_STAGING_ROOT, user_dir)
        if not os.path.isdir(upath):
            continue
        for upload_dir in os.listdir(upath):
            dpath = os.path.join(upath, upload_dir)
            try:
                if os.path.isdir(dpath) and os.path.getmtime(dpath) < cutoff:
                    shutil.rmtree(dpath, ignore_errors=True)
            except OSError:
                pass


@router.post("/upload-media-header/init")
async def init_chunked_media_upload(payload: dict, user: dict = Depends(get_current_user)):
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    """CR-036 B.3 (E-B3-2): start a chunked media upload session."""
    mime = payload.get("mime", "")
    total_size = int(payload.get("total_size", 0))
    total_chunks = int(payload.get("total_chunks", 0))
    kind = _classify_mime(mime)
    if not kind:
        raise HTTPException(status_code=400, detail=f"Unsupported media type: {mime}")
    cap = _MEDIA_CAPS[kind]["max_bytes"]
    if total_size <= 0 or total_chunks <= 0:
        raise HTTPException(status_code=400, detail="total_size and total_chunks must be positive")
    if total_size > cap:
        max_mb = cap // (1024 * 1024)
        raise HTTPException(status_code=413, detail=f"File too large. Max for {kind}: {max_mb} MB.")

    _sweep_stale_staging()
    upload_id = str(uuid.uuid4())
    sdir = _staging_dir(user["id"], upload_id)
    os.makedirs(sdir, exist_ok=True)
    manifest = {
        "user_id": user["id"],
        "filename": payload.get("filename", "media"),
        "mime": mime,
        "total_size": total_size,
        "total_chunks": total_chunks,
        "template_slug": payload.get("template_slug", "template"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(os.path.join(sdir, "manifest.json"), "w") as f:
        json.dump(manifest, f)
    return {"upload_id": upload_id, "chunk_size": _MEDIA_CHUNK_SIZE}


def _load_manifest(user_id: str, upload_id: str) -> tuple:
    sdir = _staging_dir(user_id, upload_id)
    mpath = os.path.join(sdir, "manifest.json")
    if not os.path.isfile(mpath):
        raise HTTPException(status_code=404, detail="Upload session not found or expired")
    with open(mpath) as f:
        manifest = json.load(f)
    if manifest.get("user_id") != user_id:
        raise HTTPException(status_code=404, detail="Upload session not found or expired")
    return sdir, manifest


@router.post("/upload-media-header/chunk/{upload_id}")
async def upload_media_chunk(
    upload_id: str,
    chunk_index: int = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """CR-036 B.3 (E-B3-3): receive one chunk of a chunked media upload."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    sdir, manifest = _load_manifest(user["id"], upload_id)
    if chunk_index < 0 or chunk_index >= manifest["total_chunks"]:
        raise HTTPException(status_code=400, detail=f"chunk_index out of range 0..{manifest['total_chunks'] - 1}")
    data = await file.read()
    with open(os.path.join(sdir, f"part_{chunk_index:05d}"), "wb") as f:
        f.write(data)
    received = len([p for p in os.listdir(sdir) if p.startswith("part_")])
    return {"received": received, "total": manifest["total_chunks"]}


@router.post("/upload-media-header/complete/{upload_id}")
async def complete_chunked_media_upload(upload_id: str, user: dict = Depends(get_current_user)):
    """CR-036 B.3 (E-B3-4): assemble chunks, validate, run Meta + S3 processing."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    sdir, manifest = _load_manifest(user["id"], upload_id)
    missing = [
        i for i in range(manifest["total_chunks"])
        if not os.path.isfile(os.path.join(sdir, f"part_{i:05d}"))
    ]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing chunk indices: {missing}")

    contents = b""
    for i in range(manifest["total_chunks"]):
        with open(os.path.join(sdir, f"part_{i:05d}"), "rb") as f:
            contents += f.read()

    try:
        if len(contents) != manifest["total_size"]:
            raise HTTPException(
                status_code=400,
                detail=f"Assembled size {len(contents)} does not match declared total_size {manifest['total_size']}",
            )
        result = await _process_media_upload(
            user, contents, manifest["mime"], manifest["filename"], manifest["template_slug"]
        )
    finally:
        shutil.rmtree(sdir, ignore_errors=True)
    return result


@router.get("/custom-templates")
async def list_custom_templates(user: dict = Depends(get_current_user)):
    """List all custom templates for the user."""
    templates = await db.custom_templates.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(None)
    return {"templates": templates}


@router.put("/custom-templates/{template_id}")
async def update_custom_template(template_id: str, payload: dict, user: dict = Depends(get_current_user)):
    """Update a custom template. Sets status back to 'draft' on edit."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    import re
    now = datetime.now(timezone.utc).isoformat()

    # CR-036 B.1 Q16: block content edits on approved templates, but allow media re-upload
    existing = await db.custom_templates.find_one(
        {"id": template_id, "user_id": user["id"]}, {"status": 1}
    )
    if existing and existing.get("status") == "approved":
        has_media_update = payload.get("header_handle") or payload.get("send_media_url")
        if has_media_update:
            media_update = {
                "header_handle": payload.get("header_handle") or None,
                "send_media_url": payload.get("send_media_url") or None,
                "send_media_filename": payload.get("send_media_filename") or None,
                "header_media_mime": payload.get("header_media_mime") or None,
                "media_url": payload.get("media_url") or payload.get("send_media_url") or None,
                "needs_media_reupload": False,
                "updated_at": now,
            }
            await db.custom_templates.update_one(
                {"id": template_id, "user_id": user["id"]},
                {"$set": media_update}
            )
            return {"message": "Media updated on approved template", "id": template_id}
        else:
            raise HTTPException(
                status_code=400,
                detail="Cannot edit an approved template. Clone and create a new version instead.",
            )

    body = payload.get("body", "")
    variables = list(set(re.findall(r'\{\{\d+\}\}', body)))
    variables.sort(key=lambda v: int(v.strip('{}') or 0))
    
    update_fields = {
        "template_name": payload.get("template_name", "").strip(),
        "category": payload.get("category", "utility"),
        "language": payload.get("language", "en"),
        "header_type": payload.get("header_type", "none"),
        "header_content": payload.get("header_content", ""),
        "body": body,
        "footer": payload.get("footer", ""),
        "buttons": payload.get("buttons", []),
        "media_url": payload.get("media_url", ""),
        # CR-036 B.1: persist media fields on update
        "header_handle": payload.get("header_handle") or None,
        "send_media_url": payload.get("send_media_url") or None,
        "send_media_filename": payload.get("send_media_filename") or None,
        "header_media_mime": payload.get("header_media_mime") or None,
        "needs_media_reupload": False,
        "variables": variables,
        "status": "draft",
        "updated_at": now
    }
    
    result = await db.custom_templates.update_one(
        {"id": template_id, "user_id": user["id"]},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template updated", "id": template_id}


@router.get("/templates-in-use")
async def get_templates_in_use(user: dict = Depends(get_current_user)):
    """Return set of template_ids that are in use (event maps + campaigns)."""
    in_use = set()
    async for m in db.whatsapp_event_template_map.find(
        {"user_id": user["id"]}, {"_id": 0, "template_id": 1}
    ):
        in_use.add(str(m.get("template_id", "")))
    async for c in db.campaigns.find(
        {"user_id": user["id"]}, {"_id": 0, "template_id": 1}
    ):
        in_use.add(str(c.get("template_id", "")))
    return {"in_use_template_ids": list(in_use)}


@router.delete("/custom-templates/{template_id}")
async def delete_custom_template(template_id: str, user: dict = Depends(get_current_user)):
    """Delete a custom template. Blocked if template is in use (event mapping or campaign)."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    # Rule 2: Check if template is in use
    event_usage = await db.whatsapp_event_template_map.find_one(
        {"user_id": user["id"], "template_id": template_id}
    )
    if event_usage:
        raise HTTPException(
            status_code=400,
            detail=f"Template is mapped to event '{event_usage.get('event_key')}' and cannot be deleted. Unmap it first."
        )
    campaign_usage = await db.campaigns.find_one(
        {"user_id": user["id"], "template_id": template_id}
    )
    if campaign_usage:
        raise HTTPException(
            status_code=400,
            detail=f"Template is used in campaign '{campaign_usage.get('name')}' and cannot be deleted."
        )
    result = await db.custom_templates.delete_one(
        {"id": template_id, "user_id": user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    # Clean up variable mappings for this template
    await db.whatsapp_template_variable_map.delete_one(
        {"user_id": user["id"], "template_id": template_id}
    )
    return {"message": "Template deleted"}


@router.patch("/custom-templates/{template_id}/labels")
async def save_template_labels(template_id: str, payload: dict, user: dict = Depends(get_current_user)):
    """
    CR-DIRECT-SEND: Save variable labels for a template so external servers can
    send flat JSON payloads (e.g. {"name": "Rahul", "meeting_link": "..."}).
    Payload: {"variable_labels": {"1": "name", "2": "meeting_link"}}
    """
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    labels = payload.get("variable_labels", {})
    # Normalize: keys to strings
    normalized = {str(k): str(v).strip() for k, v in labels.items() if str(v).strip()}
    now = datetime.now(timezone.utc).isoformat()
    result = await db.custom_templates.update_one(
        {"id": template_id, "user_id": user["id"]},
        {"$set": {"variable_labels": normalized, "updated_at": now}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Labels saved", "template_id": template_id, "variable_labels": normalized}


@router.put("/custom-templates/{template_id}/submit")
async def submit_custom_template(template_id: str, user: dict = Depends(get_current_user)):
    """Submit a draft template for approval (changes status to pending)."""
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    now = datetime.now(timezone.utc).isoformat()
    result = await db.custom_templates.update_one(
        {"id": template_id, "user_id": user["id"], "status": "draft"},
        {"$set": {"status": "pending", "updated_at": now}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found or not in draft status")
    return {"message": "Template submitted for approval", "id": template_id}


@router.get("/custom-templates/{template_id}/status")
async def check_template_status(template_id: str, user: dict = Depends(get_current_user)):
    """Check template approval status from Meta and update local record."""
    template = await db.custom_templates.find_one(
        {"id": template_id, "user_id": user["id"]}, {"_id": 0}
    )
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    meta_tid = template.get("meta_template_id")
    if not meta_tid:
        return {"status": template.get("status", "draft"), "meta_status": None}

    user_doc = await db.users.find_one(
        {"id": user["id"]}, {"meta_waba_id": 1, "meta_access_token": 1}
    )
    waba_id = user_doc.get("meta_waba_id") if user_doc else None
    access_token = user_doc.get("meta_access_token") if user_doc else None

    if not waba_id or not access_token:
        return {"status": template.get("status", "draft"), "meta_status": "credentials_missing"}

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{os.environ['META_GRAPH_API_URL']}/{meta_tid}",
                headers={"Authorization": f"Bearer {access_token}"},
                params={"fields": "name,status,category,language,quality_score,rejected_reason"}
            )
        meta_data = resp.json()
        meta_status = meta_data.get("status", "").upper()

        status_map = {"APPROVED": "approved", "REJECTED": "rejected", "PENDING": "pending", "IN_APPEAL": "pending", "PAUSED": "approved"}
        new_status = status_map.get(meta_status, template.get("status", "pending"))

        update_fields = {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}
        reject_reason = meta_data.get("rejected_reason") or (meta_data.get("quality_score") or {}).get("reasons")
        if reject_reason:
            update_fields["reject_reason"] = str(reject_reason)

        await db.custom_templates.update_one({"id": template_id}, {"$set": update_fields})

        return {
            "status": new_status,
            "meta_status": meta_status,
            "meta_template_id": meta_tid,
            "reject_reason": reject_reason,
        }
    except Exception as e:
        logging.warning(f"Meta status check failed for template {template_id}: {e}")
        return {"status": template.get("status", "pending"), "meta_status": "check_failed", "error": str(e)}


@router.get("/check-template-name")
async def check_template_name(name: str = Query(...), user: dict = Depends(get_current_user)):
    """Check if template name already exists on Meta WABA."""
    user_doc = await db.users.find_one(
        {"id": user["id"]}, {"meta_waba_id": 1, "meta_access_token": 1}
    )
    waba_id = user_doc.get("meta_waba_id") if user_doc else None
    access_token = user_doc.get("meta_access_token") if user_doc else None

    if not waba_id or not access_token:
        return {"exists": False, "error": "credentials_missing"}

    clean_name = name.strip().lower().replace(" ", "_")
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{os.environ['META_GRAPH_API_URL']}/{waba_id}/message_templates",
                headers={"Authorization": f"Bearer {access_token}"},
                params={"name": clean_name, "fields": "name,status", "limit": 5}
            )
        data = resp.json()
        templates = data.get("data", [])
        exists = any(t.get("name") == clean_name for t in templates)
        return {"exists": exists, "clean_name": clean_name}
    except Exception:
        return {"exists": False, "error": "check_failed"}



# ---- Meta API Integration (Stage 1) ----

@router.post("/meta/create-template")
async def create_meta_template(payload: dict, user: dict = Depends(get_current_user)):
    """
    Stage 1: Create a WhatsApp template on Meta using Graph API.
    Transforms our template format to Meta's required format.
    """
    # CR-061 revised: backend authoring gates removed — all tenants can create templates.
    # _require_crm_templates_enabled(user)   # CR-061 (removed: owner wants all tenants to author)
    # Get user's Meta credentials
    user_doc = await db.users.find_one(
        {"id": user["id"]}, 
        {"meta_waba_id": 1, "meta_access_token": 1}
    )
    
    if not user_doc:
        raise HTTPException(status_code=400, detail="User not found")

    # BUG-QA-04: Validation runs BEFORE Meta credential check so users get
    # meaningful input errors regardless of whether WABA is configured.
    # Build Meta API payload fields (also needed later for Meta API call)
    template_name = payload.get("template_name", "").strip().lower().replace(" ", "_")
    category = payload.get("category", "utility").upper()
    language = payload.get("language", "en")

    # ---- V1-V4 Backend safety-net validation ----
    import re as _re
    body_text_raw = payload.get("body", "")
    footer_raw = payload.get("footer", "")
    header_content_raw = payload.get("header_content", "")
    header_type_raw = payload.get("header_type", "none")
    validation_errors = []

    # V1: Single-brace detection
    single_brace_re = _re.compile(r'(?<!\{)\{(\d+)\}(?!\})')
    if single_brace_re.search(body_text_raw):
        validation_errors.append("Body contains single-brace variables like {1}. Use double braces: {{1}}, {{2}}")
    if single_brace_re.search(header_content_raw):
        validation_errors.append("Header contains single-brace variables. Use {{1}} instead of {1}")
    if single_brace_re.search(footer_raw):
        validation_errors.append("Footer contains single-brace variables. Variables are not allowed in footer")

    # V2: Sequential variable numbering in body
    body_var_nums = sorted(set(int(m) for m in _re.findall(r'\{\{(\d+)\}\}', body_text_raw)))
    if body_var_nums:
        if body_var_nums[0] != 1:
            validation_errors.append("Body variables must start at {{1}}")
        expected = list(range(body_var_nums[0], body_var_nums[-1] + 1))
        if body_var_nums != expected:
            validation_errors.append(f"Body variables not sequential: {['{{' + str(n) + '}}' for n in body_var_nums]}")

    # V3: Footer cannot contain variables
    if _re.search(r'\{\{\d+\}\}', footer_raw):
        validation_errors.append("Footer cannot contain variables. Meta does not support footer variables")

    # V4: Header text — max 1 variable, must be {{1}}
    if header_type_raw == "text" and header_content_raw:
        header_var_nums = [int(m) for m in _re.findall(r'\{\{(\d+)\}\}', header_content_raw)]
        if len(header_var_nums) > 1:
            validation_errors.append(f"Header text allows only 1 variable ({{{{1}}}}). Found {len(header_var_nums)}")
        elif len(header_var_nums) == 1 and header_var_nums[0] != 1:
            validation_errors.append("Header variable must be {{1}}")

    # CR-036 B.1 Q18: reject {{n}} variables in media header content
    if header_type_raw in ("image", "video", "document"):
        if _re.search(r'\{\{\d+\}\}', header_content_raw):
            validation_errors.append("Dynamic variables {{n}} are not supported in media header content. Only static media headers are supported.")

    # ---- CR-066: V11-V20 Backend safety-net ----
    # V11: Unmatched formatting markers
    us_count = body_text_raw.count("_")
    if us_count % 2 != 0:
        validation_errors.append(f"Body has unmatched _ (italic marker): {us_count} found, must be even")
    tilde_count = body_text_raw.count("~")
    if tilde_count % 2 != 0:
        validation_errors.append(f"Body has unmatched ~ (strikethrough): {tilde_count} found, must be even")
    mono_count = body_text_raw.count("```")
    if mono_count % 2 != 0:
        validation_errors.append(f"Body has unmatched ``` (monospace): {mono_count} found, must be even")
    all_stars = body_text_raw.count("*")
    bullet_stars = len(_re.findall(r'(?:^|\n)\* ', body_text_raw))
    bold_stars = all_stars - bullet_stars
    if bold_stars % 2 != 0:
        validation_errors.append(f"Body has unmatched * (bold marker): {bold_stars} non-bullet * found, must be even")

    # V12: Variable at start/end of body
    stripped_body = body_text_raw.strip()
    if _re.match(r'^\{\{\d+\}\}', stripped_body):
        validation_errors.append("Body cannot start with a variable")
    if _re.search(r'\{\{\d+\}\}$', stripped_body):
        validation_errors.append("Body cannot end with a variable")

    # V13: Adjacent variables
    if _re.search(r'\}\}\s*\{\{', body_text_raw):
        validation_errors.append("Adjacent variables without text between them")

    # V14: Formatting wrapping variables
    if _re.search(r'[*_~]\{\{\d+\}\}[*_~]', body_text_raw):
        validation_errors.append("Formatting markers wrapping variables (e.g., *{{1}}*) not allowed")

    # V15: Body hard limit
    if len(body_text_raw) > 1024:
        validation_errors.append(f"Body exceeds 1024 character limit ({len(body_text_raw)} chars)")

    # V16: Emoji count — BUG-QA-02: added missing ranges \u2B50-\u2B55, \u231A-\u231B, \u23E9-\u23F3
    emoji_re = _re.compile(r'[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF\U0001F1E0-\U0001F1FF\U00002600-\U000026FF\U00002700-\U000027BF\U00002733-\U00002734\U00002714-\U00002716\U00002764\U00002B50-\U00002B55\U0000231A-\U0000231B\U000023E9-\U000023F3]')
    emoji_count = len(emoji_re.findall(body_text_raw))
    if emoji_count > 10:
        validation_errors.append(f"Body has {emoji_count} emojis - maximum 10 allowed")

    # V20: Formatting in header/footer
    if header_type_raw == "text" and _re.search(r'[*_~`]', header_content_raw):
        validation_errors.append("Header text cannot contain formatting markers")
    if _re.search(r'[*_~`]', footer_raw):
        validation_errors.append("Footer cannot contain formatting markers")

    if validation_errors:
        raise HTTPException(status_code=400, detail=" | ".join(validation_errors))

    # Meta credentials check — only reached after all validation passes
    waba_id = user_doc.get("meta_waba_id")
    access_token = user_doc.get("meta_access_token")

    if not waba_id or not access_token:
        raise HTTPException(
            status_code=400,
            detail="Meta WABA ID and Access Token are required. Please configure them in Settings."
        )

    components = []
    
    # Header component
    header_type = payload.get("header_type", "none")
    if header_type != "none":
        header_component = {
            "type": "HEADER",
            "format": header_type.upper()
        }
        if header_type == "text":
            header_text = payload.get("header_content", "")
            header_component["text"] = header_text
            # Add example if header has variables
            header_examples = payload.get("header_examples", [])
            if "{{" in header_text and header_examples:
                header_component["example"] = {"header_text": header_examples}
        # CR-036 B.1: send Meta's opaque handle, NOT the URL (Q3, Q13 — no audio)
        elif header_type in ("image", "video", "document"):
            handle = payload.get("header_handle")
            if handle:
                header_component["example"] = {"header_handle": [handle]}
            else:
                raise HTTPException(
                    status_code=400,
                    detail="Media header template missing header_handle. Re-upload the header file first.",
                )
        components.append(header_component)
    
    # Body component (required)
    body_text = payload.get("body", "")
    if not body_text:
        raise HTTPException(status_code=400, detail="Body text is required")
    
    body_component = {
        "type": "BODY",
        "text": body_text
    }
    
    # Add body examples if variables exist
    body_examples = payload.get("body_examples", [])
    if "{{" in body_text and body_examples:
        body_component["example"] = {"body_text": [body_examples]}
    
    components.append(body_component)
    
    # Footer component (optional)
    footer_text = payload.get("footer", "")
    if footer_text:
        components.append({
            "type": "FOOTER",
            "text": footer_text
        })
    
    # Buttons component (optional)
    buttons = payload.get("buttons", [])
    if buttons:
        button_components = []
        for btn in buttons:
            btn_type = btn.get("type", "QUICK_REPLY").upper()
            btn_obj = {"type": btn_type, "text": btn.get("text", "")}
            if btn_type == "URL":
                url = btn.get("url", "")
                btn_obj["url"] = url
                # Dynamic URL: if url contains {{1}}, add example array for Meta
                if "{{1}}" in url:
                    url_example = btn.get("url_example", "")
                    if url_example:
                        btn_obj["example"] = [url_example]
            elif btn_type == "PHONE_NUMBER":
                btn_obj["phone_number"] = btn.get("phone_number", "")
            button_components.append(btn_obj)
        
        if button_components:
            components.append({
                "type": "BUTTONS",
                "buttons": button_components
            })
    
    meta_payload = {
        "name": template_name,
        "language": language,
        "category": category,
        "components": components
    }
    
    # Log payload for debugging
    import logging
    logging.info(f"Meta API payload: {meta_payload}")
    
    # Call Meta Graph API
    meta_url = f"{os.environ['META_GRAPH_API_URL']}/{waba_id}/message_templates"
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                meta_url,
                headers={
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json"
                },
                json=meta_payload
            )
        
        response_data = response.json()
        
        # Log response for debugging
        logging.info(f"Meta API response: {response.status_code} - {response_data}")
        
        if response.status_code != 200:
            error_obj = response_data.get("error", {})
            error_msg = error_obj.get("message", "Unknown error")
            error_code = error_obj.get("code", "")
            error_subcode = error_obj.get("error_subcode", "")
            user_msg = error_obj.get("error_user_msg", "")
            logging.error(f"Meta API error details: {error_obj}")
            detail_parts = [f"Meta API error: {error_msg}"]
            if user_msg:
                detail_parts.append(f"Details: {user_msg}")
            if error_code:
                detail_parts.append(f"(code: {error_code}, subcode: {error_subcode})")
            raise HTTPException(
                status_code=response.status_code, 
                detail=" | ".join(detail_parts)
            )
        
        # Save template locally with meta_template_id
        now = datetime.now(timezone.utc).isoformat()
        template_id = str(uuid.uuid4())
        
        import re
        variables = list(set(re.findall(r'\{\{\d+\}\}', body_text)))
        variables.sort(key=lambda v: int(v.strip('{}') or 0))
        
        doc = {
            "id": template_id,
            "user_id": user["id"],
            "template_name": template_name,
            "category": payload.get("category", "utility"),
            "language": language,
            "header_type": header_type,
            "header_content": payload.get("header_content", ""),
            "body": body_text,
            "footer": footer_text,
            "buttons": buttons,
            "variables": variables,
            "body_examples": body_examples,
            "header_examples": payload.get("header_examples", []),
            # BUG-010: persist media fields so campaign wizard sees has_send_media
            "media_url": payload.get("media_url") or payload.get("send_media_url") or "",
            "header_handle": payload.get("header_handle") or None,
            "send_media_url": payload.get("send_media_url") or None,
            "send_media_filename": payload.get("send_media_filename") or None,
            "header_media_mime": payload.get("header_media_mime") or None,
            "needs_media_reupload": False,
            "meta_template_id": response_data.get("id"),
            "status": "pending",  # Meta templates start as pending review
            "created_at": now,
            "updated_at": now
        }
        
        await db.custom_templates.insert_one(doc)
        doc.pop("_id", None)
        
        return {
            "message": "Template created on Meta successfully",
            "meta_template_id": response_data.get("id"),
            "template": doc
        }
        
    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Failed to connect to Meta API: {str(e)}")


# ---- AuthKey Sync API (Stage 2) ----

@router.post("/authkey/sync-templates")
async def sync_authkey_templates(user: dict = Depends(get_current_user)):
    """
    Stage 2: Sync/migrate templates from Meta to AuthKey.
    Should be called after creating template on Meta.
    CR-DIRECT-SEND: After sync, fetches AuthKey templates and saves authkey_wid
    back into matching custom_templates documents.
    """
    # Get user's AuthKey credentials
    user_doc = await db.users.find_one(
        {"id": user["id"]}, 
        {"authkey_api_key": 1, "brand_number": 1}
    )
    
    if not user_doc:
        raise HTTPException(status_code=400, detail="User not found")
    
    api_key = user_doc.get("authkey_api_key")
    brand_number = user_doc.get("brand_number")
    
    if not api_key:
        raise HTTPException(
            status_code=400, 
            detail="AuthKey API key is required. Please configure it in Settings."
        )
    
    if not brand_number:
        raise HTTPException(
            status_code=400, 
            detail="Brand number is required. Please configure it in Settings."
        )
    
    # Call AuthKey migration API
    authkey_url = os.environ['AUTHKEY_SYNC_URL']
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                authkey_url,
                headers={
                    "Authorization": f"Basic {api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "method": "migrate",
                    "brand_number": brand_number
                }
            )
        
        response_data = response.json()
        
        # Log the response for debugging
        import logging
        logging.info(f"AuthKey sync response: {response_data}")
        
        # AuthKey might return status as boolean or string, or different field names
        status = response_data.get("status") or response_data.get("Status")
        if status in [False, "false", "0", 0]:
            error_msg = response_data.get("message") or response_data.get("Message") or "Sync failed"
            raise HTTPException(status_code=400, detail=f"AuthKey sync error: {error_msg}")

        # CR-DIRECT-SEND: After successful sync, fetch all AuthKey templates and
        # back-fill authkey_wid on matching custom_templates docs (match by name).
        wid_updates = 0
        try:
            async with httpx.AsyncClient(timeout=20) as client:
                tpl_resp = await client.post(
                    os.environ['AUTHKEY_TEMPLATES_URL'],
                    headers={"Authorization": f"Basic {api_key}", "Content-Type": "application/json"},
                    json={"channel": "whatsapp"},
                )
            tpl_data = tpl_resp.json()
            authkey_templates = tpl_data.get("data", [])

            # Build lookup: normalized_name -> wid
            authkey_by_name = {}
            for at in authkey_templates:
                norm = (at.get("temp_name") or "").strip().lower().replace(" ", "_")
                if norm and at.get("wid"):
                    authkey_by_name[norm] = str(at["wid"])

            # Match against local custom_templates
            local_templates = await db.custom_templates.find(
                {"user_id": user["id"]}, {"_id": 0, "id": 1, "template_name": 1, "status": 1}
            ).to_list(None)
            for ct in local_templates:
                norm_ct = (ct.get("template_name") or "").strip().lower().replace(" ", "_")
                wid = authkey_by_name.get(norm_ct)
                if wid:
                    # CR-037: Preserve authoritative "rejected" status set by
                    # Meta status check (Path A at check_template_status).
                    # Only overwrite to "approved" when current status is not
                    # already "rejected". authkey_wid is always back-filled so
                    # that AuthKey WID discovery remains functional even on
                    # rejected templates (owner may fix + resubmit).
                    current_status = ct.get("status", "draft")
                    update_set = {"authkey_wid": wid}
                    if current_status != "rejected":
                        update_set["status"] = "approved"
                    await db.custom_templates.update_one(
                        {"id": ct["id"]},
                        {"$set": update_set}
                    )
                    wid_updates += 1

            # CR-073: Import externally-created AuthKey templates into custom_templates.
            # Only runs when tenant has Meta WABA configured (Q2=b).
            imported_count = 0
            user_meta = await db.users.find_one(
                {"id": user["id"]},
                {"_id": 0, "meta_waba_id": 1, "meta_access_token": 1}
            )
            meta_waba_id      = (user_meta or {}).get("meta_waba_id")
            meta_access_token = (user_meta or {}).get("meta_access_token")

            if meta_waba_id and meta_access_token:
                # Build sets of already-known wids and names to avoid duplicates
                local_wids  = {str(c.get("authkey_wid", "")) for c in local_templates if c.get("authkey_wid")}
                local_names = {(ct.get("template_name") or "").strip().lower() for ct in local_templates}
                now_import  = datetime.now(timezone.utc).isoformat()
                meta_api_url = os.environ.get("META_GRAPH_API_URL", "https://graph.facebook.com/v18.0")

                for at in authkey_templates:
                    wid_str  = str(at.get("wid", ""))
                    tpl_name = (at.get("temp_name") or "").strip()

                    # Skip if already in custom_templates (idempotent guard)
                    if wid_str in local_wids or tpl_name.lower() in local_names:
                        continue

                    try:
                        # Fetch button data from Meta
                        buttons = []
                        async with httpx.AsyncClient(timeout=10) as meta_client:
                            meta_resp = await meta_client.get(
                                f"{meta_api_url}/{meta_waba_id}/message_templates",
                                params={
                                    "name": tpl_name,
                                    "fields": "name,status,components",
                                    "access_token": meta_access_token,
                                }
                            )
                        for meta_tpl in meta_resp.json().get("data", []):
                            for comp in meta_tpl.get("components", []):
                                if comp.get("type") != "BUTTONS":
                                    continue
                                for btn in comp.get("buttons", []):
                                    btn_type = btn.get("type", "")
                                    btn_obj  = {"type": btn_type, "text": btn.get("text", "")}
                                    if btn_type == "URL":
                                        url = btn.get("url", "")
                                        btn_obj["url"] = url
                                        if "{{1}}" in url:
                                            btn_obj["url_type"] = "dynamic"
                                            url_base = url.split("{{")[0]
                                            btn_obj["url_base"] = url_base
                                            # Q1=a: use Meta's example; strip base if full URL returned
                                            example_arr = btn.get("example") or []
                                            raw_ex = example_arr[0] if example_arr else ""
                                            if raw_ex.startswith(url_base):
                                                btn_obj["url_example"] = raw_ex[len(url_base):]
                                            else:
                                                btn_obj["url_example"] = raw_ex
                                        else:
                                            btn_obj["url_type"] = "static"
                                    elif btn_type == "PHONE_NUMBER":
                                        btn_obj["phone_number"] = btn.get("phone_number", "")
                                    buttons.append(btn_obj)

                        # Extract body variables
                        body_text = at.get("temp_body", "")
                        variables = sorted(
                            set(re.findall(r'\{\{\d+\}\}', body_text)),
                            key=lambda v: int(v.strip("{}") or 0)
                        )

                        # Build and insert new custom_templates document
                        doc = {
                            "id": str(uuid.uuid4()),
                            "user_id": user["id"],
                            "template_name": tpl_name,
                            "category": (at.get("temp_category") or "utility").lower(),
                            "language": at.get("temp_language") or "en",
                            "header_type": "none",
                            "header_content": "",
                            "body": body_text,
                            "footer": "",
                            "buttons": buttons,
                            "variables": variables,
                            "body_examples": [],
                            "header_examples": [],
                            "media_url": "",
                            "header_handle": None,
                            "send_media_url": None,
                            "send_media_filename": None,
                            "header_media_mime": None,
                            "needs_media_reupload": False,
                            "meta_template_id": None,
                            "authkey_wid": wid_str,
                            "status": "approved" if at.get("temp_status") == 1 else "pending",
                            "created_at": now_import,
                            "updated_at": now_import,
                        }
                        await db.custom_templates.insert_one(doc)
                        imported_count += 1
                        logging.info(f"CR-073: imported '{tpl_name}' wid={wid_str} user={user['id']}")

                    except Exception as import_err:
                        logging.warning(f"CR-073: skipped '{tpl_name}' wid={wid_str}: {import_err}")
                        continue

        except Exception as wid_err:
            logging.warning(f"CR-DIRECT-SEND: could not back-fill authkey_wid after sync: {wid_err}")

        return {
            "message": "Templates synced to AuthKey successfully",
            "response": response_data,
            "wid_updates": wid_updates,
            "imported_count": imported_count if 'imported_count' in locals() else 0,
        }
        
    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Failed to connect to AuthKey API: {str(e)}")


# ---- Combined Create & Sync ----

@router.post("/create-and-sync-template")
async def create_and_sync_template(payload: dict, user: dict = Depends(get_current_user)):
    """
    Combined endpoint: Creates template on Meta (Stage 1) then syncs to AuthKey (Stage 2).
    """
    # Stage 1: Create on Meta
    meta_result = await create_meta_template(payload, user)
    
    # Stage 2: Sync to AuthKey
    try:
        sync_result = await sync_authkey_templates(user)
        return {
            "message": "Template created and synced successfully",
            "meta_result": meta_result,
            "sync_result": sync_result
        }
    except HTTPException as e:
        # Meta succeeded but AuthKey failed - return partial success
        return {
            "message": "Template created on Meta but AuthKey sync failed",
            "meta_result": meta_result,
            "sync_error": e.detail
        }


@router.put("/event-template-map")
async def save_event_template_map(payload: dict, user: dict = Depends(get_current_user)):
    """Save event→template mappings. Upserts per (user_id, event_key)."""
    mappings = payload.get("mappings", [])
    if not mappings:
        raise HTTPException(status_code=400, detail="No mappings provided.")
    now = datetime.now(timezone.utc).isoformat()
    saved = 0
    for m in mappings:
        event_key = m.get("event_key")
        template_id = m.get("template_id")
        template_name = m.get("template_name", "")
        if not event_key or template_id is None:
            continue
        await db.whatsapp_event_template_map.update_one(
            {"user_id": user["id"], "event_key": event_key},
            {"$set": {
                "template_id": template_id,
                "template_name": template_name,
                "is_enabled": m.get("is_enabled", True),
                "updated_at": now,
            }, "$setOnInsert": {
                "user_id": user["id"],
                "event_key": event_key,
                "created_at": now,
            }},
            upsert=True,
        )
        saved += 1
    return {"message": "Mappings saved", "count": saved}


@router.post("/event-template-map/{event_key}/toggle")
async def toggle_event_mapping(event_key: str, user: dict = Depends(get_current_user)):
    """Toggle is_enabled for an event mapping."""
    doc = await db.whatsapp_event_template_map.find_one(
        {"user_id": user["id"], "event_key": event_key}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Mapping not found")
    new_val = not doc.get("is_enabled", True)
    await db.whatsapp_event_template_map.update_one(
        {"user_id": user["id"], "event_key": event_key},
        {"$set": {"is_enabled": new_val, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"event_key": event_key, "is_enabled": new_val}


@router.get("/event-template-map")
async def get_event_template_map(user: dict = Depends(get_current_user)):
    """Get saved event→template mappings for the current user."""
    docs = await db.whatsapp_event_template_map.find(
        {"user_id": user["id"]}, {"_id": 0, "user_id": 0}
    ).to_list(100)
    return {"mappings": docs}

@router.delete("/event-template-map/{event_key}")
async def delete_event_template_map(event_key: str, user: dict = Depends(get_current_user)):
    """Delete/unmap a template from an event."""
    result = await db.whatsapp_event_template_map.delete_one(
        {"user_id": user["id"], "event_key": event_key}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event mapping not found")
    return {"message": "Template unmapped successfully", "event_key": event_key}

@router.get("/template-variable-map")
async def get_template_variable_mappings(user: dict = Depends(get_current_user)):
    """Get all template variable mappings for the current user."""
    docs = await db.whatsapp_template_variable_map.find(
        {"user_id": user["id"]}, {"_id": 0, "user_id": 0}
    ).to_list(100)
    return {"mappings": docs}

@router.put("/template-variable-map/{template_id}")
async def save_template_variable_mapping(
    template_id: str,
    data: dict,
    user: dict = Depends(get_current_user)
):
    """Save variable mappings for a template + return warnings for incompatible event/variable combos."""
    from core.whatsapp_variables import fills_on, COUPON_VARIABLE_KEYS, VARIABLES_BY_KEY

    now = datetime.now(timezone.utc).isoformat()
    clean_mappings = {k: v for k, v in (data.get("mappings") or {}).items() if v and v != "none"}
    modes = data.get("modes") or {}

    # CR-004 P2.5-B: Validate coupon_pick mode entries
    for placeholder, mapped_value in clean_mappings.items():
        if modes.get(placeholder) != "coupon_pick":
            continue
        parts = mapped_value.split(":")
        if len(parts) != 3 or parts[0] != "coupon":
            raise HTTPException(
                400,
                f"Invalid coupon_pick format for {placeholder}: expected 'coupon:<id>:<field>'"
            )
        coupon_id, field = parts[1], parts[2]
        if ":" in coupon_id:
            raise HTTPException(
                400,
                f"Invalid coupon_id for {placeholder}: must not contain ':'"
            )
        if field not in ("code", "title", "discount", "expiry"):
            raise HTTPException(
                400,
                f"Invalid coupon field '{field}' for {placeholder}: must be code|title|discount|expiry"
            )
        cpn = await db.coupons.find_one(
            {"id": coupon_id, "user_id": user["id"]}, {"_id": 1}
        )
        if not cpn:
            raise HTTPException(
                404,
                f"Coupon '{coupon_id}' not found or does not belong to your account"
            )

    # CR-020: Validate menu_pick mode entries
    for placeholder, mapped_value in clean_mappings.items():
        if modes.get(placeholder) != "menu_pick":
            continue
        parts = mapped_value.split(":")
        if len(parts) != 3 or parts[0] not in ("menu_item", "menu_category"):
            raise HTTPException(
                400,
                f"Invalid menu_pick format for {placeholder}: expected 'menu_item:<id>:<field>' or 'menu_category:<id>:<field>'"
            )
        entity_type, entity_id, field = parts[0], parts[1], parts[2]
        if entity_type == "menu_item" and field not in ("name", "price"):
            raise HTTPException(400, f"Invalid menu_item field '{field}' for {placeholder}: must be name|price")
        if entity_type == "menu_category" and field not in ("name",):
            raise HTTPException(400, f"Invalid menu_category field '{field}' for {placeholder}: must be name")

    # CR-015 T6 (2026-05-29): Validate map-mode var_keys against registry.
    # Block save if any map-mode mapping uses an unknown variable key.
    map_mode_errors = []
    for placeholder, mapped_value in clean_mappings.items():
        mode = modes.get(placeholder, "map")
        if mode in ("text", "coupon_pick", "menu_pick"):
            continue  # text = literal string (valid); coupon_pick/menu_pick = already validated above
        clean_key = (mapped_value or "").strip()
        if clean_key in ("", "none"):
            continue  # explicit no-mapping, allowed
        if clean_key not in VARIABLES_BY_KEY:
            map_mode_errors.append({
                "placeholder": placeholder,
                "type": "unknown_variable",
                "message": f"Unknown variable '{mapped_value}' for {placeholder}. Pick from the available list."
            })

    if map_mode_errors:
        raise HTTPException(
            status_code=422,
            detail={"errors": map_mode_errors}
        )

    await db.whatsapp_template_variable_map.update_one(
        {"user_id": user["id"], "template_id": template_id},
        {"$set": {
            "user_id": user["id"],
            "template_id": template_id,
            "template_name": data.get("template_name", ""),
            "mappings": clean_mappings,
            "modes": modes,
            "menu_pick_resolved": data.get("menu_pick_resolved") or {},
            "updated_at": now,
        }},
        upsert=True,
    )

    # P2: Compute warnings — check each map-mode variable against events using this template
    warnings = []

    # CR-015 T6: Warn on text-mode values that look like placeholders/notes
    suspicious_tokens = ("missing", "todo", "tbd", "n/a", "none", "placeholder", "test")
    for placeholder, mapped_value in clean_mappings.items():
        mode = modes.get(placeholder, "map")
        if mode != "text":
            continue
        val_lower = (mapped_value or "").lower().strip()
        is_suspicious = (
            any(token in val_lower for token in suspicious_tokens)
            or (mapped_value or "").strip() != (mapped_value or "")  # trailing/leading whitespace
        )
        if is_suspicious:
            warnings.append({
                "placeholder": placeholder,
                "type": "text_mode_suspicious_value",
                "variable": mapped_value,
                "message": f"{placeholder}: '{mapped_value}' looks like a placeholder — this text will be sent to customers literally."
            })

    event_mappings = await db.whatsapp_event_template_map.find(
        {"user_id": user["id"], "template_id": template_id},
        {"_id": 0, "event_key": 1},
    ).to_list(50)

    for em in event_mappings:
        event_key = em.get("event_key")
        if not event_key:
            continue
        for placeholder, var_key in clean_mappings.items():
            if modes.get(placeholder) in ("text", "coupon_pick", "menu_pick"):
                continue
            if not fills_on(var_key, event_key):
                warnings.append({
                    "event": event_key,
                    "placeholder": placeholder,
                    "variable": var_key,
                    "message": f"Variable '{var_key}' does not reliably fill on event '{event_key}'.",
                })

    return {
        "message": "Variable mappings saved",
        "template_id": template_id,
        "mappings": clean_mappings,
        "warnings": warnings,
    }


@router.post("/test-template")
async def test_template(request: TestTemplateRequest, user: dict = Depends(get_current_user)):
    """
    Send a test WhatsApp message using the specified template.
    Used to verify template configuration before enabling automation.
    """
    # Get user's AuthKey API key
    user_doc = await db.users.find_one({"id": user["id"]}, {"authkey_api_key": 1})
    api_key = user_doc.get("authkey_api_key") if user_doc else None
    
    if not api_key:
        raise HTTPException(
            status_code=400,
            detail="AuthKey API key not configured. Please add your API key in Settings."
        )
    
    # Validate phone number
    phone = request.phone.replace(" ", "").replace("-", "")
    if not phone or len(phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")

    # CR-036 B.1 Q17: auto-inject stored send_media_url for media templates
    effective_media_url = request.media_url
    effective_media_filename = request.media_filename
    if not effective_media_url and request.template_id:
        _tpl_lookup = await db.custom_templates.find_one(
            {"user_id": user["id"], "authkey_wid": request.template_id},
            {"header_type": 1, "send_media_url": 1, "send_media_filename": 1},
        )
        if not _tpl_lookup:
            _tpl_lookup = await db.custom_templates.find_one(
                {"user_id": user["id"], "id": request.template_id},
                {"header_type": 1, "send_media_url": 1, "send_media_filename": 1},
            )
        if _tpl_lookup and _tpl_lookup.get("header_type") in ("image", "video", "document"):
            if _tpl_lookup.get("send_media_url"):
                effective_media_url = _tpl_lookup["send_media_url"]
                effective_media_filename = _tpl_lookup.get("send_media_filename") or "file"
            else:
                raise HTTPException(
                    status_code=400,
                    detail="Media template missing send_media_url — re-upload the header file first.",
                )
    
    # Build message
    message = WhatsAppMessage(
        phone=phone,
        country_code=request.country_code.replace("+", ""),
        template_id=request.template_id,
        body_values=request.body_values,
        media_url=effective_media_url,
        media_filename=effective_media_filename,
        customer_id=None,
        button_values=request.button_values or None,  # CR-069
    )
    
    # Send test message
    result = await send_single_message(api_key, message)

    # CR-004 P3.5 Commit 4 (G11): unify Path B with Path A — use the canonical
    # log_message_attempt writer so test sends share one row schema with real
    # event sends. is_test=True flags these for dashboard exclusion.
    await log_message_attempt(
        db,
        user["id"],
        None,                                            # customer_id
        phone,                                           # customer_phone
        "test",                                          # event_type
        request.template_id,
        result,
        template_name=None,
        campaign_id=None,
        country_code=request.country_code.replace("+", ""),
        body_values=request.body_values,
        customer_name=None,
        reference_type=None,
        reference_id=None,
        pos_order_id=None,
        idempotency_key=None,                            # test sends are NOT deduped
        is_test=True,
        media_url=request.media_url,
        media_filename=request.media_filename,
        message_body_text=None,
        channel="wp",
    )

    if result.success:
        return {
            "success": True,
            "message_id": result.message_id,
            "message": f"Test message sent successfully to +{request.country_code} {phone}",
            "response_data": result.response_data
        }
    else:
        return {
            "success": False,
            "error": result.error,
            "message": f"Failed to send test message: {result.error}"
        }


# ============================================
# Message Status Dashboard APIs
# ============================================

@router.get("/message-stats")
async def get_message_stats(
    user: dict = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_test: bool = False,
):
    """Get message statistics by status.

    CR-004 P3.5 Commit 6: `include_test` (default False) excludes is_test=True
    rows so owner-initiated /test-template sends don't inflate the stats.
    """
    query = {"user_id": user["id"]}

    if not include_test:
        query["is_test"] = {"$ne": True}

    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to
        else:
            query["created_at"] = {"$lte": date_to}

    # Aggregate counts by status
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    results = await db.whatsapp_message_logs.aggregate(pipeline).to_list(10)

    stats = {
        "total": 0,
        "delivered": 0,
        "read": 0,
        "pending": 0,
        "rejected": 0,
        # CR-036 B.2 (E-B2-2): expose failed bucket + media_missing sub-count
        "failed": 0,
        "media_missing": 0,
    }

    for r in results:
        status = r["_id"]
        count = r["count"]
        stats["total"] += count
        if status in stats:
            stats[status] = count

    # CR-036 B.2 (E-B2-2): media_missing sub-count honours same query
    # (include_test + date range) as the status aggregation above.
    stats["media_missing"] = await db.whatsapp_message_logs.count_documents(
        {**query, "status_note": "media_missing"}
    )

    return stats


# CR-042 + BUG-009: shared query builder for message-logs list + export.
# Extracted from the original inline construction in get_message_logs so that
# GET /message-logs and GET /message-logs/export produce byte-identical filter
# semantics against the whatsapp_message_logs collection.
def _build_message_log_query(
    user_id: str,
    status: Optional[str] = None,
    event_type: Optional[str] = None,
    campaign_id: Optional[str] = None,
    run_id: Optional[str] = None,          # CR-042 + BUG-009: new dimension
    template_name: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_test: bool = False,
    status_note: Optional[str] = None,     # CR-036 B.2 (E-B2-3): media_missing chip
) -> Dict[str, Any]:
    query: Dict[str, Any] = {"user_id": user_id}

    if not include_test:
        query["is_test"] = {"$ne": True}

    if status and status != "all":
        query["status"] = status
    if event_type and event_type != "all":
        query["event_type"] = event_type
    if campaign_id and campaign_id != "all":
        # BUG-006 fix: match campaign_id OR reference_id for backward compat
        # (old logs stored run_id in campaign_id, campaign_id in reference_id)
        if "$and" not in query:
            query["$and"] = []
        query["$and"].append({"$or": [
            {"campaign_id": campaign_id},
            {"reference_id": campaign_id},
        ]})
    if run_id and run_id != "all":
        # CR-042 + BUG-009: same $or shape as campaign_id — captures legacy
        # logs (which stored run_id in campaign_id) and current logs
        # (reference_id = run_id). Combined with campaign_id filter, this
        # narrows to the exact run of the exact campaign.
        if "$and" not in query:
            query["$and"] = []
        query["$and"].append({"$or": [
            {"campaign_id": run_id},
            {"reference_id": run_id},
        ]})
    if template_name and template_name != "all":
        query["template_name"] = template_name
    # CR-036 B.2 (E-B2-3): status_note dimension (currently: "media_missing")
    if status_note and status_note != "all":
        query["status_note"] = status_note
    if search:
        # CR-004 P3.5 Commit 6: regex-escape + match across name AND phone
        safe = re.escape(search.strip())
        if safe:
            if "$and" not in query:
                query["$and"] = []
            query["$and"].append({"$or": [
                {"customer_phone": {"$regex": safe, "$options": "i"}},
                {"customer_name": {"$regex": safe, "$options": "i"}},
            ]})
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to
        else:
            query["created_at"] = {"$lte": date_to}
    return query


@router.get("/message-logs")
async def get_message_logs(
    user: dict = Depends(get_current_user),
    status: Optional[str] = None,
    event_type: Optional[str] = None,
    campaign_id: Optional[str] = None,
    run_id: Optional[str] = None,          # CR-042 + BUG-009: new filter dimension
    template_name: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_test: bool = False,
    status_note: Optional[str] = None,     # CR-036 B.2 (E-B2-3): media_missing chip
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200)
):
    """Get paginated message logs with filters.

    CR-004 P3.5 Commit 6:
      - `include_test` (default False) excludes is_test=True rows.
      - `search` matches BOTH customer_phone and customer_name (regex-escaped).

    CR-042 + BUG-009: adds `run_id` filter — narrows to messages of a specific
    campaign run via the same $or pattern used for campaign_id (BUG-006 compat).
    """
    query = _build_message_log_query(
        user_id=user["id"],
        status=status, event_type=event_type,
        campaign_id=campaign_id, run_id=run_id,
        template_name=template_name, search=search,
        date_from=date_from, date_to=date_to,
        include_test=include_test,
        status_note=status_note,           # CR-036 B.2 (E-B2-3)
    )

    # Get total count
    total = await db.whatsapp_message_logs.count_documents(query)

    # Get logs
    logs = await db.whatsapp_message_logs.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    return {
        "total": total,
        "logs": logs,
        "skip": skip,
        "limit": limit
    }


# CR-042: Message report download — CSV + XLSX with 5000-row cap.
# Reuses _build_message_log_query() so filter semantics stay identical to the
# list endpoint. Filter-aware (MessageStatusPage export button) + run-scoped
# (CampaignHistoryPage per-row export button) share this single endpoint.
_EXPORT_HEADERS = [
    ("Sent At", "created_at"),
    ("Phone", "customer_phone"),
    ("Name", "customer_name"),
    ("Event / Campaign", "_event_or_campaign"),
    ("Template", "template_name"),
    ("Status", "status"),
    ("Delivered At", "delivered_at"),
    ("Read At", "read_at"),
    ("Rejected At", "rejected_at"),
    ("Error Reason", "failure_reason"),
    ("Message ID", "message_id"),
    ("Test Send", "is_test"),
    ("Resend Count", "resend_count"),      # CR-065 (Q-F=a)
    ("Last Resend At", "last_resend_at"),  # CR-065 (Q-F=a)
]
_EXPORT_ROW_CAP = 5000
_EXPORT_BRAND_COLOR = "F26B33"  # match CR-035 XLSX styling


def _resolve_event_or_campaign(log: dict) -> str:
    """Human-readable dimension for the report column."""
    if log.get("event_type"):
        return log["event_type"]
    if log.get("campaign_name"):
        return log["campaign_name"]
    if log.get("campaign_id"):
        return f"campaign:{log['campaign_id']}"
    return ""


def _row_from_log(log: dict) -> List[str]:
    row = []
    for _, key in _EXPORT_HEADERS:
        if key == "_event_or_campaign":
            row.append(_resolve_event_or_campaign(log))
        elif key == "is_test":
            row.append("Yes" if log.get("is_test") else "No")
        else:
            v = log.get(key, "")
            row.append("" if v is None else str(v))
    return row


@router.get("/message-logs/export")
async def export_message_logs(
    user: dict = Depends(get_current_user),
    format: str = "csv",                          # 'csv' or 'xlsx'
    status: Optional[str] = None,
    event_type: Optional[str] = None,
    campaign_id: Optional[str] = None,
    run_id: Optional[str] = None,
    template_name: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_test: bool = False,
    status_note: Optional[str] = None,     # CR-036 B.2 (E-B2-3): media_missing chip
):
    """CR-042: Stream WhatsApp message logs as CSV or XLSX.

    Honours the same filter dimensions as /message-logs. Caps output at
    _EXPORT_ROW_CAP rows (5000, owner-locked) — surfaces via X-Row-Count and
    X-Row-Cap response headers so frontend can toast when the cap is reached.
    """
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")

    query = _build_message_log_query(
        user_id=user["id"],
        status=status, event_type=event_type,
        campaign_id=campaign_id, run_id=run_id,
        template_name=template_name, search=search,
        date_from=date_from, date_to=date_to,
        include_test=include_test,
        status_note=status_note,           # CR-036 B.2 (E-B2-3)
    )
    cursor = db.whatsapp_message_logs.find(
        query, {"_id": 0}
    ).sort("created_at", -1).limit(_EXPORT_ROW_CAP)
    logs = await cursor.to_list(length=_EXPORT_ROW_CAP)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    tenant_slug = (user.get("business_name") or user.get("id") or "tenant")
    tenant_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", tenant_slug)[:32]
    filename_base = f"message_report_{tenant_slug}_{ts}"

    common_headers = {
        "X-Row-Count": str(len(logs)),
        "X-Row-Cap": str(_EXPORT_ROW_CAP),
    }

    if format == "csv":
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow([h for h, _ in _EXPORT_HEADERS])
        for log in logs:
            writer.writerow(_row_from_log(log))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_base}.csv"',
                **common_headers,
            },
        )

    # xlsx branch
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Message Report"
    ws.append([h for h, _ in _EXPORT_HEADERS])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_EXPORT_BRAND_COLOR)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for log in logs:
        ws.append(_row_from_log(log))
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_base}.xlsx"',
            **common_headers,
        },
    )


@router.get("/message-filters")
async def get_message_filters(user: dict = Depends(get_current_user)):
    """Get available filter options for message logs from master data"""
    
    # Get events from master list (POS + CRM events)
    event_types = AUTOMATION_EVENTS
    
    # Get templates from authkey templates + custom templates
    # CR-004 P3.5 Commit 6: dedupe by normalized key (lowercase + strip) to collapse
    # casing/whitespace variants while preserving the first-seen original case.
    template_names: Dict[str, str] = {}

    def _add_template(name: Optional[str]):
        if not name:
            return
        key = name.strip().lower()
        if key and key not in template_names:
            template_names[key] = name

    # Get AuthKey templates if API key configured
    if user.get("authkey_api_key"):
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    os.environ['AUTHKEY_TEMPLATES_URL'],
                    params={
                        "authkey": user["authkey_api_key"],
                    },
                    timeout=10
                )
                if response.status_code == 200:
                    data = response.json()
                    templates = data.get("templates", [])
                    for t in templates:
                        _add_template(t.get("temp_name") or t.get("name"))
        except Exception:
            pass

    # Get custom templates from DB
    custom_templates = await db.custom_templates.find(
        {"user_id": user["id"]}, {"_id": 0, "name": 1}
    ).to_list(100)
    for t in custom_templates:
        _add_template(t.get("name"))

    # Get configured template mappings
    mappings = await db.whatsapp_event_template_map.find(
        {"user_id": user["id"]}, {"_id": 0, "template_name": 1}
    ).to_list(100)
    for m in mappings:
        _add_template(m.get("template_name"))

    # Get campaigns
    campaigns = await db.campaigns.find(
        {"user_id": user["id"]}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)

    return {
        "statuses": MESSAGE_STATUSES,
        "event_types": event_types,
        "template_names": sorted(template_names.values()),
        "campaigns": campaigns
    }


@router.post("/status-callback")
async def message_status_callback(request: Request):
    """
    AuthKey delivery-report webhook. Public endpoint (no auth — called by AuthKey).

    CR-004 P3.5 Commit 5: locked payload schema. Audit-first design:
      1. Capture raw body + headers into whatsapp_callback_logs BEFORE parsing.
      2. Optional HMAC verification (dormant until AUTHKEY_WEBHOOK_SECRET env var set).
      3. Defensive id extraction — AuthKey canonical key is `logid` (lowercase).
      4. Status translation table (locked from real sample 2026-05-28).
      5. `time` (Asia/Kolkata local string) parsed into delivered_at/read_at/rejected_at.
      6. State machine guards out-of-order events (e.g. late `delivered` after `read`).
      7. Recipient sanity check (mobile vs country_code+customer_phone).
      8. Always pushes to status_history (even on ignored transitions, for audit).
    """
    received_at = datetime.now(timezone.utc).isoformat()

    # ---- 1. Capture raw body FIRST (audit-first) ----
    try:
        raw_bytes = await request.body()
    except Exception:
        raw_bytes = b""

    # CR-004 P3.5 hotfix (2026-05-28): AuthKey delivery webhooks arrive as
    # application/x-www-form-urlencoded on the wire (real captured sample),
    # not JSON as the original sample suggested. Support both formats; pick
    # parser by Content-Type, fall back to "try JSON then form".
    content_type = (request.headers.get("content-type") or "").lower()
    payload: Dict[str, Any] = {}

    def _parse_form(b: bytes) -> Dict[str, Any]:
        try:
            decoded = b.decode("utf-8", errors="replace")
            parsed = parse_qs(decoded, keep_blank_values=True)
            # Flatten single-value lists; keep lists for repeated keys.
            return {k: (v[0] if len(v) == 1 else v) for k, v in parsed.items()}
        except Exception:
            return {}

    def _parse_json(b: bytes) -> Dict[str, Any]:
        try:
            obj = json.loads(b) if b else {}
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    if "application/x-www-form-urlencoded" in content_type:
        payload = _parse_form(raw_bytes)
        if not payload:
            payload = _parse_json(raw_bytes)
    elif "application/json" in content_type:
        payload = _parse_json(raw_bytes)
        if not payload:
            payload = _parse_form(raw_bytes)
    else:
        # Unknown content-type: be defensive — try JSON first, then form.
        payload = _parse_json(raw_bytes) or _parse_form(raw_bytes)

    callback_log = {
        "id": str(uuid.uuid4()),
        "received_at": received_at,
        "headers": {k.lower(): v for k, v in request.headers.items()},
        "raw_body": raw_bytes.decode("utf-8", errors="replace") if raw_bytes else "",
        "parsed": payload,
        "logid": payload.get("logid") or payload.get("LogID") or payload.get("log_id"),
        "verdict": "pending",
        "verdict_reason": None,
    }

    async def _persist_callback_and_return(verdict: str, reason: Optional[str], http_resp: Dict[str, Any]):
        callback_log["verdict"] = verdict
        callback_log["verdict_reason"] = reason
        try:
            await db.whatsapp_callback_logs.insert_one(callback_log)
        except Exception as exc:
            logger.error(f"Failed to persist whatsapp_callback_log: {exc}")
        return http_resp

    # ---- 2. HMAC verification (dormant until B2 — secret in env) ----
    secret = os.environ.get("AUTHKEY_WEBHOOK_SECRET")
    if secret:
        sig_header = (
            request.headers.get("x-auth-signature")
            or request.headers.get("X-Auth-Signature")
            or ""
        )
        expected = hmac.new(secret.encode("utf-8"), raw_bytes, hashlib.sha256).hexdigest()
        if not sig_header or not hmac.compare_digest(expected, sig_header):
            return await _persist_callback_and_return(
                "rejected_signature",
                f"sig_header={sig_header!r} expected_hex_sha256_mismatch",
                {"success": False, "error": "Invalid signature"},
            )

    # ---- 3. Defensive id extraction (logid canonical, others as fallback) ----
    logid = (
        payload.get("logid")
        or payload.get("LogID")
        or payload.get("log_id")
        or payload.get("message_id")
        or payload.get("msgId")
    )
    if not logid:
        return await _persist_callback_and_return(
            "rejected_no_logid",
            None,
            {"success": False, "error": "logid required"},
        )

    # ---- 4. Status translation (locked map) ----
    raw_status = (payload.get("status") or "").lower()
    status_map = {
        "sent": "pending",
        "delivered": "delivered",
        "read": "read",
        "failed": "rejected",
        "undelivered": "rejected",
        "rejected": "rejected",
        "not sent": "rejected",
        "not_sent": "rejected",
        "expired": "rejected",
    }
    mapped_status = status_map.get(raw_status)
    if not mapped_status:
        return await _persist_callback_and_return(
            "unknown_status",
            f"status={raw_status!r}",
            {"success": False, "error": f"unknown status: {raw_status}"},
        )

    # ---- 5. Time parsing (IST -> UTC); preserve raw verbatim ----
    time_raw = payload.get("time")
    ts_utc_iso = received_at
    if time_raw:
        try:
            ts_local = datetime.strptime(time_raw, "%Y-%m-%d %H:%M:%S").replace(
                tzinfo=ZoneInfo("Asia/Kolkata")
            )
            ts_utc_iso = ts_local.astimezone(timezone.utc).isoformat()
        except Exception:
            logger.warning(
                f"webhook time parse failed: {time_raw!r}, falling back to received_at"
            )

    # ---- 6. Row lookup (CR-039: composite (message_id, customer_phone) with fallback) ----
    # AuthKey may return duplicate LogIDs for concurrent single-mobile sends, so
    # message_id alone is not unique. Try composite lookup first to disambiguate;
    # fall back to message_id-only for legacy rows without matching mobile.
    webhook_mobile = str(payload.get("mobile") or "")
    row = None
    if webhook_mobile and len(webhook_mobile) >= 10:
        # AuthKey sends mobile with country code (e.g. 919035133228);
        # CRM stores customer_phone as last 10 digits (e.g. 9035133228).
        mobile_last_10 = webhook_mobile[-10:]
        row = await db.whatsapp_message_logs.find_one(
            {"message_id": logid, "customer_phone": mobile_last_10},
            {"_id": 0},
        )
    if not row:
        row = await db.whatsapp_message_logs.find_one({"message_id": logid}, {"_id": 0})
    if not row:
        return await _persist_callback_and_return(
            "no_matching_row",
            f"logid={logid}",
            {"success": True, "logid": logid, "updated": False},
        )

    # ---- 7. State machine ----
    new_status = next_status(row.get("status"), mapped_status)

    # ---- 8. Recipient sanity check ----
    expected_mobile = f"{row.get('country_code', '')}{row.get('customer_phone', '')}"
    mobile_mismatch = bool(webhook_mobile and expected_mobile and webhook_mobile != expected_mobile)

    # ---- 9. Build $set update ----
    set_fields: Dict[str, Any] = {"updated_at": received_at, "time_raw": time_raw}

    if payload.get("meta_messageid"):
        set_fields["meta_message_id"] = payload["meta_messageid"]
    if payload.get("keypress") is not None:
        set_fields["keypress"] = payload["keypress"]
    if payload.get("button_param_value"):
        set_fields["button_param_value"] = payload["button_param_value"]
    if payload.get("channel"):
        set_fields["channel"] = payload["channel"]
    if mobile_mismatch:
        # CR-039: Composite lookup already attempted above. A persistent mismatch
        # means the correct row cannot be identified — refuse to update rather
        # than corrupt data on the wrong row.
        logger.warning(
            f"CR-039 webhook mobile mismatch after composite lookup: "
            f"payload={webhook_mobile!r} row={expected_mobile!r} logid={logid}"
        )
        return await _persist_callback_and_return(
            "ambiguous_row",
            f"mobile_mismatch payload={webhook_mobile} row={expected_mobile}",
            {"success": True, "logid": logid, "updated": False},
        )

    # Dispatch time -> status-specific timestamp field (CR-041)
    # Timestamps and failure_reason are the AUTHORITATIVE record of when the
    # transition ACTUALLY happened. On a state-machine-rejected duplicate/late
    # webhook (transition_ignored), leave them untouched to preserve the
    # original event time. Only apply when the state machine allowed the
    # transition. Metadata fields above (meta_message_id, keypress,
    # button_param_value, channel) remain ungated because they may legitimately
    # arrive on a late webhook (e.g. button-press response) even if the status
    # transition itself is blocked.

    # ---- 10. Apply status only if transition is valid ----
    if new_status:
        set_fields["status"] = new_status
        applied = True
    else:
        applied = False

    if applied:
        if mapped_status == "delivered":
            set_fields["delivered_at"] = ts_utc_iso
        elif mapped_status == "read":
            set_fields["read_at"] = ts_utc_iso
        elif mapped_status == "rejected":
            set_fields["rejected_at"] = ts_utc_iso
            set_fields["failure_reason"] = (
                payload.get("reason")
                or payload.get("Reason")
                or payload.get("error")
                or payload.get("Error")
                or payload.get("description")
                or payload.get("Message")
                or payload.get("message")
                or raw_status
            )

    # ---- 11. Always push to status_history (audit) ----
    history_entry = {
        "status": mapped_status,
        "timestamp": ts_utc_iso,
        "received_at": received_at,
        "action": "webhook",
        "applied": applied,
        "raw_payload": payload,
    }

    try:
        await db.whatsapp_message_logs.update_one(
            {"id": row["id"]},
            {"$set": set_fields, "$push": {"status_history": history_entry}},
        )
    except Exception as exc:
        logger.exception(f"Failed to update whatsapp_message_logs row id={row.get('id')}: {exc}")
        return await _persist_callback_and_return(
            "db_update_failed",
            str(exc),
            {"success": False, "error": "internal_error"},
        )

    return await _persist_callback_and_return(
        "applied" if applied else "transition_ignored",
        None if applied else f"{row.get('status')!r}->{mapped_status!r} not allowed",
        {
            "success": True,
            "logid": logid,
            "status": new_status or row.get("status"),
            "applied": applied,
        },
    )


class ResendRequest(BaseModel):
    message_ids: List[str]


@router.post("/resend")
async def resend_messages(
    request: ResendRequest,
    user: dict = Depends(get_current_user)
):
    """Resend failed/pending messages"""
    if not request.message_ids:
        raise HTTPException(status_code=400, detail="No message IDs provided")
    
    # Get messages to resend
    # CR-036 B.3 (E-B3-5, Q20=a): "failed" (media_missing) rows are now eligible
    messages = await db.whatsapp_message_logs.find({
        "id": {"$in": request.message_ids},
        "user_id": user["id"],
        "status": {"$in": ["pending", "rejected", "failed"]}
    }, {"_id": 0}).to_list(len(request.message_ids))
    
    if not messages:
        raise HTTPException(status_code=404, detail="No eligible messages found")
    
    # Get user's AuthKey API key
    authkey_api_key = user.get("authkey_api_key")
    if not authkey_api_key:
        raise HTTPException(status_code=400, detail="WhatsApp API key not configured")
    
    results = []
    now = datetime.now(timezone.utc).isoformat()
    now_dt = datetime.now(timezone.utc)
    GRACE_MINUTES = 30  # CR-004 P3.5 Commit 6: in-flight grace window
    _tpl_media_cache = {}  # CR-036 B.3 (E-B3-5): one template lookup per template_id

    async def _lookup_template_media(template_id: str):
        if template_id in _tpl_media_cache:
            return _tpl_media_cache[template_id]
        proj = {"send_media_url": 1, "send_media_filename": 1}
        tpl = await db.custom_templates.find_one(
            {"authkey_wid": str(template_id), "user_id": user["id"]}, proj
        ) or await db.custom_templates.find_one(
            {"id": template_id, "user_id": user["id"]}, proj
        ) or await db.templates.find_one(
            {"id": template_id, "user_id": user["id"]}, proj
        )
        _tpl_media_cache[template_id] = tpl
        return tpl

    for msg in messages:
        # CR-004 P3.5 Commit 6: guard against resending in-flight pending messages.
        # If a message is `pending`, was sent in the last 30 minutes, AND has only
        # the initial_send entry in status_history, the delivery webhook just
        # hasn't arrived yet. Resending would duplicate-deliver to the customer.
        try:
            if msg.get("status") == "pending":
                created_at_raw = msg.get("created_at")
                history_len = len(msg.get("status_history") or [])
                if created_at_raw and history_len <= 1:
                    try:
                        created_dt = datetime.fromisoformat(created_at_raw.replace("Z", "+00:00"))
                    except Exception:
                        created_dt = None
                    if created_dt and (now_dt - created_dt) < timedelta(minutes=GRACE_MINUTES):
                        results.append({
                            "id": msg["id"],
                            "success": False,
                            "skipped": True,
                            "error": "in_flight_grace_period",
                            "message": f"Pending < {GRACE_MINUTES} min, awaiting delivery report",
                        })
                        continue

            # CR-036 B.3 (E-B3-5, Q20=a): media_missing rows re-check template media
            _resend_media_url = None
            _resend_media_filename = None
            if msg.get("status") == "failed":
                if msg.get("status_note") != "media_missing":
                    results.append({
                        "id": msg["id"],
                        "success": False,
                        "skipped": True,
                        "error": "not_resendable",
                        "message": "Only media_missing failed rows can be resent",
                    })
                    continue
                tpl = await _lookup_template_media(msg.get("template_id"))
                if not tpl or not tpl.get("send_media_url"):
                    results.append({
                        "id": msg["id"],
                        "success": False,
                        "skipped": True,
                        "error": "media_still_missing",
                        "message": "Template still has no uploaded media — re-upload first",
                    })
                    continue
                _resend_media_url = tpl["send_media_url"]
                _resend_media_filename = tpl.get("send_media_filename")

            # Build WhatsApp message
            wa_msg = WhatsAppMessage(
                phone=msg.get("customer_phone"),
                country_code=msg.get("country_code", "91"),
                template_id=msg.get("template_id"),
                body_values=msg.get("body_values", {}),
                media_url=_resend_media_url,
                media_filename=_resend_media_filename
            )

            # Send message
            result = await send_single_message(authkey_api_key, wa_msg)

            new_status = "pending" if result.success else "rejected"
            new_message_id = result.message_id if result.success else msg.get("message_id")

            # CR-036 B.3 (E-B3-5): clear media_missing note once resend succeeds
            _set_fields = {
                "message_id": new_message_id,
                "status": new_status,
                "updated_at": now,
                "resend_count": msg.get("resend_count", 0) + 1,
                "last_resend_at": now
            }
            if result.success and msg.get("status_note") == "media_missing":
                _set_fields["status_note"] = None

            # Update the log with new attempt
            await db.whatsapp_message_logs.update_one(
                {"id": msg["id"]},
                {
                    "$set": _set_fields,
                    "$push": {
                        "status_history": {
                            "status": new_status,
                            "timestamp": now,
                            "action": "resend",
                            "success": result.success,
                            "error": result.error if not result.success else None
                        }
                    }
                }
            )

            results.append({
                "id": msg["id"],
                "success": result.success,
                "new_status": new_status,
                "error": result.error if not result.success else None
            })

        except Exception as e:
            results.append({
                "id": msg["id"],
                "success": False,
                "error": str(e)
            })

    success_count = sum(1 for r in results if r.get("success"))
    skipped_count = sum(1 for r in results if r.get("skipped"))

    return {
        "total": len(messages),
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": len(messages) - success_count - skipped_count,
        "results": results
    }
