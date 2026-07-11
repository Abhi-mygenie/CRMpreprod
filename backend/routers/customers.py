from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, Request, UploadFile, File
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import uuid
import os
import httpx
import json
import logging
import csv
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger("customer_sync")

from core.database import db
from core.auth import get_current_user
from core.helpers import generate_qr_code, build_customer_query, _coerce_pos_id, _pos_id_query_variants
from core.loyalty import calculate_tier as _calc_tier
from models.schemas import (
    Customer, CustomerCreate, CustomerUpdate,
    Segment, SegmentCreate, SegmentUpdate,
    ImportLog, ImportPreviewResponse, ImportPreviewRow, ImportRowError
)

router = APIRouter(prefix="/customers", tags=["Customers"])

# In-memory customer sync status tracking (kept as fast-read cache; F9 persists to DB)
customer_sync_status = {}

# CR-035: Ordered list of (csv_header, customer_dict_key) for export
EXPORT_FIELDS = [
    ("Name",            "name"),
    ("Phone",           "phone"),
    ("Email",           "email"),
    ("Date of Birth",   "dob"),
    ("Anniversary",     "anniversary"),
    ("Gender",          "gender"),
    ("City",            "city"),
    ("Address",         "address"),
    ("State",           "state"),
    ("Pincode",         "pincode"),
    ("Total Points",    "total_points"),
    ("Tier",            "tier"),
    ("Wallet Balance",  "wallet_balance"),
    ("Total Visits",    "total_visits"),
    ("Total Spent",     "total_spent"),
    ("Last Visit",      "last_visit"),
    ("Tags",            "tags"),
    ("WhatsApp Opt-in", "whatsapp_opt_in"),
    ("VIP",             "vip_flag"),
    ("Lead Source",     "lead_source"),
    ("Customer Type",   "customer_type"),
    ("Created At",      "created_at"),
]


def _parse_import_file(content: bytes, filename: str) -> list:
    """CR-035: Parse CSV or Excel file bytes into list of row dicts with 1-based row index."""
    rows = []
    fname = (filename or "").lower()
    try:
        if fname.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                raise HTTPException(status_code=400, detail="Excel file is empty")
            headers = [str(h).strip().lower() if h else "" for h in raw_rows[0]]
            for i, row in enumerate(raw_rows[1:], start=2):
                row_dict = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
                row_dict["_row"] = i
                rows.append(row_dict)
        else:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for i, row in enumerate(reader, start=2):
                row_dict = {k.strip().lower(): v.strip() for k, v in row.items()}
                row_dict["_row"] = i
                rows.append(row_dict)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {str(e)}")
    return rows


def _validate_and_classify_row(row: dict, existing_phones: set) -> dict:
    """CR-035: Validate a single parsed row. Returns status: new|update|error."""
    row_num = row.get("_row", 0)
    name  = row.get("name", "").strip()
    phone = row.get("phone", "").strip()

    if not name:
        return {"status": "error", "row": row_num, "reason": "Missing name"}
    if not phone:
        return {"status": "error", "row": row_num, "reason": "Missing phone number"}

    clean_phone = phone.replace(" ", "").replace("-", "").lstrip("+")
    if clean_phone.startswith("91") and len(clean_phone) == 12:
        clean_phone = clean_phone[2:]
    if not clean_phone.isdigit():
        return {"status": "error", "row": row_num, "reason": f"Invalid phone format: '{phone}'"}
    if len(clean_phone) != 10:
        return {"status": "error", "row": row_num, "reason": f"Phone must be 10 digits, got {len(clean_phone)}"}

    status = "update" if clean_phone in existing_phones else "new"
    raw_tags = row.get("tags", "").strip()
    tags_list = [t.strip() for t in raw_tags.split(",") if t.strip()] if raw_tags else []

    return {
        "status":  status,
        "row":     row_num,
        "name":    name,
        "phone":   clean_phone,
        "email":   row.get("email", "").strip() or None,
        "dob":     row.get("dob", "").strip() or None,
        "city":    row.get("city", "").strip() or None,
        "address": row.get("address", "").strip() or None,
        "tags":    tags_list,
        "reason":  None,
    }


# CR-001B-fix Phase 2A F9 helpers — persistent migration_sync_logs (customer sync side)
async def _cust_log_progress(log_id: str, fields: dict):
    """Persist progress fields to migration_sync_logs. Best-effort, never raises."""
    try:
        await db.migration_sync_logs.update_one({"id": log_id}, {"$set": fields})
    except Exception:
        logger.exception("migration_sync_logs update_failed log_id=%s", log_id)


async def _cust_push_failed_record(log_id: str, record: dict):
    """Append a failed_records entry, capped at 100. Best-effort, never raises."""
    try:
        await db.migration_sync_logs.update_one(
            {"id": log_id, "failed_records.99": {"$exists": False}},
            {"$push": {"failed_records": record}},
        )
    except Exception:
        logger.exception("migration_sync_logs push_failed_record_error log_id=%s", log_id)


async def background_customer_sync(user_id: str, mygenie_token: str):
    """Background task to sync customers"""
    mygenie_api_url = os.environ['MYGENIE_API_URL']
    
    now = datetime.now(timezone.utc).isoformat()
    synced_count = 0
    updated_count = 0
    failed_count = 0  # CR-001B-fix Phase 2A F8: per-record failure counter
    total_customers = 0
    
    # CR-001B-fix Phase 2A F9: insert persistent log row at sync start
    log_id = str(uuid.uuid4())
    log_doc = {
        "id": log_id,
        "user_id": user_id,
        "sync_type": "customer_sync",
        "status": "running",
        "started_at": now,
        "completed_at": None,
        "total_records": 0,
        "synced_count": 0,
        "updated_count": 0,
        "failed_count": 0,
        "current_page": 0,
        "total_pages": 0,
        "error": None,
        "failed_records": [],
        "created_at": now,
    }
    try:
        await db.migration_sync_logs.insert_one(log_doc)
    except Exception:
        logger.exception("migration_sync_logs insert_failed user_id=%s sync_type=customer_sync", user_id)
    
    customer_sync_status[user_id] = {
        "status": "running",
        "synced": 0,
        "updated": 0,
        "failed_count": 0,
        "total_customers": 0,
        "started_at": now,
        "error": None,
        "log_id": log_id,
        "sync_type": "customer_sync",
    }

    # ============================================================
    # CR-001C-L Phase L3 (D2 + Q-LB1 Option C, 2026-05-22)
    # ============================================================
    # D2: Block customer-sync if loyalty_settings doc is missing.
    # Q-LB1 Option C (REVISED 2026-05-23 — CR-001C-L LF-MERGE):
    #   The "Loyalty Program" master toggle (`loyalty_enabled`) now drives
    #   BOTH realtime POS earning AND migration clean-slate recompute.
    #   The previous hidden flag `loyalty_clean_slate_recalc` is deprecated
    #   and no longer read.
    #   When loyalty_enabled=True  → hard-init counters to 0, drop synthetic
    #                                 backfill, use allow-list $set on
    #                                 existing customer (C11 safety).
    #   When loyalty_enabled=False → legacy behavior preserved verbatim
    #                                 (trust MyGenie aggregates).
    loyalty_settings_doc = await db.loyalty_settings.find_one(
        {"user_id": user_id}, {"_id": 0}
    )
    if not loyalty_settings_doc:
        err_msg = (
            "Migration blocked: loyalty_settings doc not found for this "
            "restaurant. Configure Loyalty Settings (master toggle + tier "
            "thresholds) before triggering customer sync. See CR-001C-L "
            "blueprint D2."
        )
        customer_sync_status[user_id]["status"] = "failed"
        customer_sync_status[user_id]["error"] = err_msg
        await _cust_log_progress(log_id, {
            "status": "failed",
            "error": err_msg,
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
        return

    # CR-001C-L LF-MERGE (2026-05-23): clean_slate now derives from loyalty_enabled.
    clean_slate = bool(loyalty_settings_doc.get("loyalty_enabled", False))
    
    try:
        async with httpx.AsyncClient() as client:
            page = 1
            last_page = 1
            customer_index = 0

            while page <= last_page:
                response = await client.post(
                    f"{mygenie_api_url}/api/v1/vendoremployee/whatsappcrm/customer-migration?page={page}",
                    headers={
                        "Authorization": f"Bearer {mygenie_token}",
                        "Content-Type": "application/json; charset=UTF-8",
                        "X-localization": "en"
                    },
                    json={},
                    timeout=60.0
                )

                if response.status_code != 200:
                    customer_sync_status[user_id]["status"] = "failed"
                    customer_sync_status[user_id]["error"] = f"API error on page {page}: {response.status_code}"
                    # CR-001B-fix Phase 2A F9: persist failure
                    await _cust_log_progress(log_id, {
                        "status": "failed",
                        "error": f"API error on page {page}: {response.status_code}",
                        "completed_at": datetime.now(timezone.utc).isoformat(),
                        "synced_count": synced_count,
                        "updated_count": updated_count,
                        "failed_count": failed_count,
                    })
                    return

                data = response.json()
                customer_list = data.get("customers", [])
                last_page = data.get("last_page", 1)
                total_customers = data.get("total_customers", len(customer_list))

                # Update progress
                customer_sync_status[user_id]["total_customers"] = total_customers
                customer_sync_status[user_id]["current_page"] = page
                customer_sync_status[user_id]["total_pages"] = last_page

                # Store total from POS in user record (only on first page)
                if page == 1:
                    await db.users.update_one(
                        {"id": user_id},
                        {"$set": {"total_customers_in_pos": total_customers}}
                    )

                for i, mygenie_customer in enumerate(customer_list):
                  # CR-001B-fix Phase 2A F8: per-record try/except so one bad customer doesn't kill the loop
                  try:
                    # CR-001B-fix Phase 2B F6: explicit null-safety on pos_customer_id.
                    # Without this we'd later raise KeyError on mygenie_customer["id"] inside customer_data.
                    raw_pos_customer_id = mygenie_customer.get("id")
                    if not raw_pos_customer_id:
                        failed_count += 1
                        logger.warning(
                            "customer_sync skipping record with missing id user_id=%s page=%s idx=%s payload_keys=%s",
                            user_id, page, i, list(mygenie_customer.keys())[:10],
                        )
                        await _cust_push_failed_record(log_id, {
                            "pos_id": None,
                            "page": page,
                            "index": i,
                            "error": "missing id field in MyGenie customer payload",
                            "ts": datetime.now(timezone.utc).isoformat(),
                        })
                        customer_index += 1
                        continue
                    # CR-001B-fix Phase 2B F5: normalise pos_customer_id to str for consistent storage
                    pos_customer_id_str = _coerce_pos_id(raw_pos_customer_id)

                    # DIAGNOSTIC: log every record we attempt so we can pinpoint the crashing one
                    logger.info(
                        "customer_sync user_id=%s page=%s idx=%s pos_customer_id=%s name=%r phone=%r",
                        user_id, page, i,
                        pos_customer_id_str,
                        (mygenie_customer.get("name") or "")[:40],
                        mygenie_customer.get("phone"),
                    )
                    customer_data = {
                        "user_id": user_id,
                        "name": mygenie_customer.get("name") or "Unknown",
                        "phone": mygenie_customer.get("phone") or "",
                        "country_code": mygenie_customer.get("country_code") or "+91",
                        "email": mygenie_customer.get("email") or f"customer{pos_customer_id_str}@mygenie.local",
                        "dob": mygenie_customer.get("dob"),
                        "anniversary": mygenie_customer.get("anniversary"),
                        "gst_name": mygenie_customer.get("gst_name"),
                        "gst_number": mygenie_customer.get("gst_number"),
                        # CR-001C-L Phase L3 (C2 + C10-mig, 2026-05-22):
                        #   clean_slate=True  → hard-init counters to 0 (DO NOT trust
                        #     MyGenie loyalty_point/total_points_earned/redeemed/wallet/
                        #     coupon aggregates; order-sync recomputes from scratch).
                        #   clean_slate=False → legacy behavior preserved (trust MyGenie
                        #     aggregates as before).
                        "total_points": 0 if clean_slate else mygenie_customer.get("loyalty_point", 0),
                        "total_points_earned": 0 if clean_slate else int(mygenie_customer.get("total_points_earned") or 0),
                        "total_points_redeemed": 0 if clean_slate else int(mygenie_customer.get("total_points_redeemed") or 0),
                        "wallet_balance": 0.0 if clean_slate else float(mygenie_customer.get("wallet_balance") or 0),
                        "total_wallet_received": 0.0 if clean_slate else float(mygenie_customer.get("total_wallet_received") or 0),
                        "total_wallet_used": 0.0 if clean_slate else float(mygenie_customer.get("total_wallet_used") or 0),
                        "total_coupon_used": 0 if clean_slate else mygenie_customer.get("total_coupon_used", 0),
                        # CR-001B-fix Phase 2B F5: store pos_customer_id as str
                        "pos_customer_id": pos_customer_id_str,
                        "pos_id": mygenie_customer.get("pos_id"),
                        "pos_restaurant_id": mygenie_customer.get("restaurant_id"),
                        "mygenie_synced": True,
                        "last_synced_at": now,
                        "last_updated_at": mygenie_customer.get("updated_time"),
                    }

                    # Map customer_addresses from MyGenie into CRM addresses[] format
                    mygenie_addresses = mygenie_customer.get("customer_addresses", [])
                    crm_addresses = []
                    for idx, mg_addr in enumerate(mygenie_addresses):
                        if not isinstance(mg_addr, dict):
                            logger.warning(
                                "customer_sync skipping non-dict address user_id=%s pos_customer_id=%s idx=%s value=%r",
                                user_id, pos_customer_id_str, idx, mg_addr,
                            )
                            continue
                        crm_addr = {
                            "id": f"addr_{uuid.uuid4().hex[:12]}",
                            "pos_address_id": str(mg_addr.get("id", "")),
                            "address_type": mg_addr.get("address_type") or "Other",
                            "address": mg_addr.get("address") or "",
                            "house": mg_addr.get("house") or None,
                            "floor": mg_addr.get("floor") or None,
                            "road": mg_addr.get("road") or None,
                            "city": mg_addr.get("city") or None,
                            "state": None,
                            "pincode": mg_addr.get("pincode") or None,
                            "country": "India",
                            "latitude": mg_addr.get("latitude") or None,
                            "longitude": mg_addr.get("longitude") or None,
                            "contact_person_name": mg_addr.get("contact_person_name") or None,
                            "contact_person_number": mg_addr.get("contact_person_number") or None,
                            "dial_code": mg_addr.get("dial_code") or None,
                            "zone_id": str(mg_addr["zone_id"]) if mg_addr.get("zone_id") is not None else None,
                            "delivery_instructions": None,
                            "is_default": idx == 0,
                            "created_at": mg_addr.get("created_at") or now,
                            "updated_at": mg_addr.get("updated_at") or now,
                        }
                        crm_addresses.append(crm_addr)
                    customer_data["addresses"] = crm_addresses

                    # CR-001C-L Phase L3 (F1, 2026-05-22): use shared tier helper
                    # instead of hardcoded ladder so thresholds match settings doc.
                    customer_data["tier"] = _calc_tier(
                        customer_data.get("total_points", 0), loyalty_settings_doc
                    )

                    # Check if exists
                    # CR-001B-fix Phase 2B F5: lookup uses $in to match both legacy int and new str
                    pos_customer_variants = _pos_id_query_variants(raw_pos_customer_id)
                    existing = await db.customers.find_one({
                        "user_id": user_id,
                        "pos_customer_id": {"$in": pos_customer_variants},
                    })

                    # CR-001B-fix Phase 3 F11: phone + country_code dedup before INSERT.
                    # Prevents NEW duplicates only. Old duplicates remain untouched (Q19).
                    # Dedup key: (user_id, phone, country_code). Per-restaurant scope unchanged.
                    if not existing and customer_data.get("phone"):
                        phone_match = await db.customers.find_one({
                            "user_id": user_id,
                            "phone": customer_data["phone"],
                            "country_code": customer_data.get("country_code", "+91"),
                        })
                        if phone_match:
                            logger.info(
                                "customer_sync F11 phone+country_code dedup matched user_id=%s phone=%s country_code=%s existing_pos_customer_id=%r incoming_pos_customer_id=%s",
                                user_id,
                                customer_data["phone"],
                                customer_data.get("country_code", "+91"),
                                phone_match.get("pos_customer_id"),
                                pos_customer_id_str,
                            )
                            existing = phone_match

                    if existing:
                        # CR-001C-L Phase L3 (C11, 2026-05-22) — re-sync safety.
                        # Under clean-slate, NEVER overwrite loyalty/wallet/coupon counters
                        # or behavioral fields (`total_visits`, `total_spent`, `last_visit`,
                        # `avg_order_value`) on re-sync. Demographics + addresses + sync
                        # metadata only. Under legacy mode, preserve current full-overwrite
                        # behavior verbatim.
                        if clean_slate:
                            _allowed_keys = {
                                "name", "phone", "country_code", "email", "dob",
                                "anniversary", "gst_name", "gst_number",
                                "pos_customer_id", "pos_id", "pos_restaurant_id",
                                "mygenie_synced", "last_synced_at", "last_updated_at",
                                "addresses",
                            }
                            safe_update = {
                                k: v for k, v in customer_data.items() if k in _allowed_keys
                            }
                            await db.customers.update_one(
                                {"id": existing["id"]},
                                {"$set": safe_update}
                            )
                        else:
                            await db.customers.update_one(
                                {"id": existing["id"]},
                                {"$set": customer_data}
                            )
                        updated_count += 1
                        customer_id = existing["id"]
                    else:
                        customer_data["id"] = str(uuid.uuid4())
                        customer_data["created_at"] = mygenie_customer.get("created_time") or now
                        customer_data["customer_type"] = mygenie_customer.get("customer_type") or "normal"
                        customer_data["notes"] = None
                        customer_data["address"] = mygenie_customer.get("address")
                        customer_data["city"] = mygenie_customer.get("city")
                        customer_data["pincode"] = mygenie_customer.get("pincode")
                        customer_data["allergies"] = []
                        customer_data["custom_field_1"] = None
                        customer_data["custom_field_2"] = None
                        customer_data["custom_field_3"] = None
                        customer_data["favorites"] = []
                        customer_data["total_visits"] = 0
                        customer_data["total_spent"] = 0.0
                        customer_data["last_visit"] = None

                        await db.customers.insert_one(customer_data)
                        synced_count += 1
                        customer_id = customer_data["id"]

                    # CR-001C-L Phase L5 (2026-05-25): synthetic historical
                    # PT/WT backfill REMOVED. Order-sync (CR-001C-L Phase L3
                    # C2) is the single source of truth for transaction
                    # history. Legacy `clean_slate=False` path no longer
                    # writes fake earn/redeem/wallet rows — it was dead post
                    # LF-MERGE since `clean_slate` now derives from
                    # `loyalty_enabled` and a restaurant with loyalty OFF
                    # has no need for synthetic loyalty history.

                    # Update progress every 10 customers
                    if (customer_index + 1) % 10 == 0:
                        customer_sync_status[user_id]["synced"] = synced_count
                        customer_sync_status[user_id]["updated"] = updated_count
                    customer_index += 1
                  except Exception as record_err:
                    # CR-001B-fix Phase 2A F8: isolate record failure, continue loop
                    failed_count += 1
                    logger.exception(
                        "customer_sync record_failed user_id=%s page=%s idx=%s pos_customer_id=%s",
                        user_id, page, i, mygenie_customer.get("id"),
                    )
                    await _cust_push_failed_record(log_id, {
                        "pos_id": str(mygenie_customer.get("id")) if mygenie_customer.get("id") is not None else None,
                        "page": page,
                        "index": i,
                        "error": str(record_err)[:500],
                        "ts": datetime.now(timezone.utc).isoformat(),
                    })
                    continue

                # Update progress after each page (in-memory + persistent F9)
                customer_sync_status[user_id]["synced"] = synced_count
                customer_sync_status[user_id]["updated"] = updated_count
                customer_sync_status[user_id]["failed_count"] = failed_count
                await _cust_log_progress(log_id, {
                    "current_page": page,
                    "total_pages": last_page,
                    "total_records": total_customers,
                    "synced_count": synced_count,
                    "updated_count": updated_count,
                    "failed_count": failed_count,
                })

                page += 1
            
            # Final update
            await db.users.update_one(
                {"id": user_id},
                {"$set": {"last_customer_sync_at": now}}
            )
            
            completed_at = datetime.now(timezone.utc).isoformat()
            customer_sync_status[user_id]["status"] = "completed"
            customer_sync_status[user_id]["synced"] = synced_count
            customer_sync_status[user_id]["updated"] = updated_count
            customer_sync_status[user_id]["completed_at"] = completed_at
            # CR-001B-fix Phase 2A F9: persist final completed state
            await _cust_log_progress(log_id, {
                "status": "completed",
                "completed_at": completed_at,
                "synced_count": synced_count,
                "updated_count": updated_count,
                "failed_count": failed_count,
            })
            
    except Exception as e:
        logger.exception(
            "customer_sync background_task_failed user_id=%s page=%s synced_so_far=%s updated_so_far=%s failed=%s",
            user_id, locals().get("page"), synced_count, updated_count, failed_count,
        )
        customer_sync_status[user_id]["status"] = "failed"
        customer_sync_status[user_id]["error"] = str(e)
        # CR-001B-fix Phase 2A F9: persist failure to DB
        await _cust_log_progress(log_id, {
            "status": "failed",
            "error": str(e)[:1000],
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "synced_count": synced_count,
            "updated_count": updated_count,
            "failed_count": failed_count,
        })


@router.post("/sync-from-mygenie")
async def sync_customers_from_mygenie(background_tasks: BackgroundTasks, request: Request, user: dict = Depends(get_current_user)):
    """
    Start background customer sync from MyGenie API.
    Returns immediately and syncs in background.
    Use /customers/sync-status to check progress.
    """
    user_id = user["id"]
    
    # Check if sync is already running
    if user_id in customer_sync_status and customer_sync_status[user_id].get("status") == "running":
        return {
            "success": False,
            "message": "Sync already in progress",
            "status": customer_sync_status[user_id]
        }
    
    # CR-008: prefer session header, fall back to DB
    mygenie_token = request.headers.get("X-MyGenie-Token")
    if not mygenie_token:
        user_record = await db.users.find_one({"id": user_id})
        mygenie_token = user_record.get("mygenie_token") if user_record else None
    
    if not mygenie_token:
        return {
            "success": False,
            "message": "MyGenie token not found. Please login again."
        }
    
    # Start background sync
    background_tasks.add_task(background_customer_sync, user_id, mygenie_token)
    
    return {
        "success": True,
        "message": "Customer sync started in background.",
        "status": "started"
    }


@router.get("/sync-status")
async def get_customer_sync_status(user: dict = Depends(get_current_user)):
    """Get current customer sync progress.

    CR-001B-fix Phase 2A F9: in-memory dict is the primary fast-read; if empty
    (e.g. after backend restart), fall back to the latest persisted entry in
    `migration_sync_logs` so the owner still sees the last known sync state.
    """
    user_id = user["id"]

    if user_id in customer_sync_status:
        return customer_sync_status[user_id]

    latest = await db.migration_sync_logs.find_one(
        {"user_id": user_id, "sync_type": "customer_sync"},
        {"_id": 0},
        sort=[("started_at", -1)],
    )
    if latest:
        return latest

    return {
        "status": "idle",
        "message": "No sync in progress"
    }


@router.post("", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, request: Request, user: dict = Depends(get_current_user)):
    # Check if phone exists for this user
    existing = await db.customers.find_one({"user_id": user["id"], "phone": customer_data.phone})
    if existing:
        raise HTTPException(status_code=400, detail="Customer with this phone already exists")
    
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Get user's MyGenie token for API sync (CR-008: header first, DB fallback)
    mygenie_token = request.headers.get("X-MyGenie-Token")
    if not mygenie_token:
        user_record = await db.users.find_one({"id": user["id"]})
        mygenie_token = user_record.get("mygenie_token") if user_record else None
    pos_customer_id = None
    
    # Sync to MyGenie if token available
    if mygenie_token:
        mygenie_api_url = os.environ['MYGENIE_API_URL']
        
        # Split name into first and last name
        name_parts = (customer_data.name or "").split(" ", 1)
        f_name = name_parts[0] if name_parts else ""
        l_name = name_parts[1] if len(name_parts) > 1 else ""
        
        mygenie_payload = {
            "phone": customer_data.phone or "",
            "f_name": f_name,
            "l_name": l_name,
            "email": customer_data.email or "",
            "gst_number": customer_data.gst_number or "",
            "gst_name": customer_data.gst_name or "",
            "date_of_birth": customer_data.dob or "",
            "date_of_anniversary": customer_data.anniversary or "",
            "membership_id": ""
        }
        
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.post(
                    f"{mygenie_api_url}/api/v1/vendoremployee/pos/user-check-create",
                    headers={
                        "Authorization": f"Bearer {mygenie_token}",
                        "Content-Type": "application/json; charset=UTF-8",
                        "X-localization": "en"
                    },
                    json=mygenie_payload,
                    timeout=15.0
                )
                
                if resp.status_code == 200:
                    mygenie_resp = resp.json()
                    pos_customer_id = mygenie_resp.get("user_id")
                    print(f"✅ Customer synced to MyGenie: {pos_customer_id}")
                else:
                    print(f"⚠️ MyGenie sync failed: {resp.status_code} - {resp.text}")
        except Exception as e:
            print(f"⚠️ MyGenie sync error (non-critical): {str(e)}")
    
    # Check for first visit bonus
    settings = await db.loyalty_settings.find_one({"user_id": user["id"]}, {"_id": 0})
    first_visit_bonus = 0
    if settings and settings.get("first_visit_bonus_enabled", False):
        first_visit_bonus = settings.get("first_visit_bonus_points", 50)
    
    customer_doc = {
        "id": customer_id,
        "user_id": user["id"],
        "created_at": now,
        "updated_at": now,
        
        # Basic Information
        "name": customer_data.name,
        "phone": customer_data.phone,
        "country_code": customer_data.country_code,
        "email": customer_data.email,
        "gender": customer_data.gender,
        "dob": customer_data.dob,
        "anniversary": customer_data.anniversary,
        "preferred_language": customer_data.preferred_language,
        "customer_type": customer_data.customer_type,
        "segment_tags": customer_data.segment_tags or [],
        
        # Contact & Marketing Permissions
        "whatsapp_opt_in": customer_data.whatsapp_opt_in,
        "whatsapp_opt_in_date": customer_data.whatsapp_opt_in_date,
        "promo_whatsapp_allowed": customer_data.promo_whatsapp_allowed,
        "promo_sms_allowed": customer_data.promo_sms_allowed,
        "email_marketing_allowed": customer_data.email_marketing_allowed,
        "call_allowed": customer_data.call_allowed,
        "is_blocked": customer_data.is_blocked,
        
        # Loyalty Information
        "total_points": first_visit_bonus,
        # CR-001C-L Phase L2 (C6, C10, 2026-05-22): defensive init of
        # lifetime earned/redeemed counters on CRM-manual-create path.
        # First-visit bonus counts toward total_points_earned per
        # Q-LOYALTY-3.
        "total_points_earned": first_visit_bonus,
        "total_points_redeemed": 0,
        "wallet_balance": 0.0,
        "tier": "Bronze",
        "referral_code": customer_data.referral_code,
        "referred_by": customer_data.referred_by,
        "membership_id": customer_data.membership_id,
        "membership_expiry": customer_data.membership_expiry,
        
        # Spending & Visit Behavior
        "total_visits": 0,
        "total_spent": 0.0,
        "avg_order_value": 0.0,
        "last_visit": None,
        "first_visit_date": now,
        "favorite_category": customer_data.favorite_category,
        "preferred_payment_mode": customer_data.preferred_payment_mode,
        
        # Customer Source & Journey
        "lead_source": customer_data.lead_source,
        "campaign_source": customer_data.campaign_source,
        "last_interaction_date": now,
        "assigned_salesperson": customer_data.assigned_salesperson,
        
        # WhatsApp CRM Tracking
        "last_whatsapp_sent": None,
        "last_whatsapp_response": None,
        "last_campaign_clicked": None,
        "last_coupon_used": None,
        "automation_status_tag": None,
        
        # Corporate Information
        "gst_name": customer_data.gst_name,
        "gst_number": customer_data.gst_number,
        "billing_address": customer_data.billing_address,
        "credit_limit": customer_data.credit_limit,
        "payment_terms": customer_data.payment_terms,
        
        # Address
        "address": customer_data.address,
        "address_line_2": customer_data.address_line_2,
        "city": customer_data.city,
        "state": customer_data.state,
        "pincode": customer_data.pincode,
        "country": customer_data.country,
        "delivery_instructions": customer_data.delivery_instructions,
        "map_location": customer_data.map_location,
        
        # Preferences
        "allergies": customer_data.allergies or [],
        "favorites": customer_data.favorites or [],
        
        # Dining Preferences
        "preferred_dining_type": customer_data.preferred_dining_type,
        "preferred_time_slot": customer_data.preferred_time_slot,
        "favorite_table": customer_data.favorite_table,
        "avg_party_size": customer_data.avg_party_size,
        "diet_preference": customer_data.diet_preference,
        "spice_level": customer_data.spice_level,
        "cuisine_preference": customer_data.cuisine_preference,
        
        # Special Occasions
        "kids_birthday": customer_data.kids_birthday or [],
        "spouse_name": customer_data.spouse_name,
        "festival_preference": customer_data.festival_preference or [],
        "special_dates": customer_data.special_dates or [],
        
        # Feedback & Flags
        "last_rating": customer_data.last_rating,
        "nps_score": customer_data.nps_score,
        "complaint_flag": customer_data.complaint_flag,
        "vip_flag": customer_data.vip_flag,
        "blacklist_flag": customer_data.blacklist_flag,
        
        # AI/Advanced
        "predicted_next_visit": customer_data.predicted_next_visit,
        "churn_risk_score": customer_data.churn_risk_score,
        "recommended_offer_type": customer_data.recommended_offer_type,
        "price_sensitivity_score": customer_data.price_sensitivity_score,
        
        # Custom Fields
        "custom_field_1": customer_data.custom_field_1,
        "custom_field_2": customer_data.custom_field_2,
        "custom_field_3": customer_data.custom_field_3,
        
        # Notes
        "notes": customer_data.notes,
        
        # POS Sync
        "pos_customer_id": pos_customer_id,
        "mygenie_synced": pos_customer_id is not None,
        "first_visit_bonus_awarded": first_visit_bonus > 0
    }
    
    await db.customers.insert_one(customer_doc)
    
    # Record first visit bonus transaction if awarded
    if first_visit_bonus > 0:
        tx_doc = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "customer_id": customer_id,
            "points": first_visit_bonus,
            "transaction_type": "bonus",
            "description": "First visit bonus - Welcome reward",
            "bill_amount": None,
            "balance_after": first_visit_bonus,
            "created_at": now
        }
        await db.points_transactions.insert_one(tx_doc)
    
    return Customer(**customer_doc)

@router.get("/sample-data")
async def get_sample_customer_data(user: dict = Depends(get_current_user)):
    """Get the first customer's data as sample for template previews."""
    customer = await db.customers.find_one(
        {"user_id": user["id"]}, {"_id": 0}
    )
    user_doc = await db.users.find_one(
        {"id": user["id"]},
        {"_id": 0, "restaurant_name": 1, "einvoice_link": 1,
         "instagram_link": 1, "google_review_link": 1, "feedback_link": 1}
    )
    restaurant_name = user_doc.get("restaurant_name", "") if user_doc else ""
    
    if not customer:
        return {"sample": {}, "restaurant_name": restaurant_name}
    
    return {
        "sample": {
            # General
            "customer_name":     customer.get("name", ""),
            "restaurant_name":   restaurant_name,
            # Loyalty
            "points_balance":    str(customer.get("total_points", 0)),
            "points_earned":     str(customer.get("total_points_earned", 0)),
            "points_redeemed":   str(customer.get("total_points_redeemed", 0)),
            "tier":              customer.get("tier", ""),
            "old_tier":          "",
            "expiring_points":   "",
            "expiry_date":       "",
            "total_visits":      str(customer.get("total_visits", 0)),
            "total_spent":       f"Rs.{customer.get('total_spent', 0)}",
            # Wallet
            "wallet_balance":    f"Rs.{customer.get('wallet_balance', 0)}",
            "amount":            f"Rs.{customer.get('total_spent', 0)}",
            # Order
            "order_id":          "",
            # CR-015a: T5 order-context sample values (static, mirror registry `example`).
            # Preview-only; live sends use build_order_event_context().
            "payment_method":      "UPI",
            "order_date":          "25 May 2026",
            "order_time":          "7:45 PM",
            "restaurant_order_id": "KM-1234",
            "transaction_id":      "TXN9876543",
            "table_id":            "T5",
            "waiter_name":         "Ramesh",
            "order_type":          "Dine-In",
            "loyalty_points_used": "200",
            "loyalty_discount":    "Rs.50",
            "wallet_used":         "Rs.100",
            "tax_amount":          "Rs.85",
            "item_count":          "3",
            "order_notes":         "No onion in biryani",
            # Coupon
            "coupon_code":       "",
            "coupon_title":      "",
            "coupon_discount":   "",
            "coupon_expiry":     "",
            # Feedback
            "rating":            "",
            # Links
            "einvoice_link":     user_doc.get("einvoice_link", "") if user_doc else "",
            "instagram_link":    user_doc.get("instagram_link", "") if user_doc else "",
            "google_review_link": user_doc.get("google_review_link", "") if user_doc else "",
            "feedback_link":     user_doc.get("feedback_link", "") if user_doc else "",
            # CR-020: Menu variable sample values for preview
            "menu_item_name":      "Veg Biryani",
            "menu_item_price":     "Rs.299",
            "menu_category_name":  "Biryani",
        },
        "restaurant_name": restaurant_name,
    }

@router.get("", response_model=List[Customer])
async def list_customers(
    search: Optional[str] = None,
    tier: Optional[str] = None,
    customer_type: Optional[str] = None,
    has_allergies: Optional[bool] = None,
    last_visit_days: Optional[int] = None,
    favorite: Optional[str] = None,
    city: Optional[str] = None,
    # New filters
    whatsapp_opt_in: Optional[str] = None,
    vip_flag: Optional[str] = None,
    diet_preference: Optional[str] = None,
    lead_source: Optional[str] = None,
    preferred_time_slot: Optional[str] = None,
    preferred_dining_type: Optional[str] = None,
    has_birthday_this_month: Optional[bool] = None,
    has_anniversary_this_month: Optional[bool] = None,
    total_visits: Optional[str] = None,
    blacklist_flag: Optional[str] = None,
    complaint_flag: Optional[str] = None,
    # Phase 3 filters
    gender: Optional[str] = None,
    total_spent: Optional[str] = None,
    is_blocked: Optional[str] = None,
    # Quick filter chips
    inactive_days: Optional[int] = None,
    most_loyal: Optional[bool] = None,
    # Feedback filter
    has_feedback: Optional[str] = None,
    # CR-043-A: filter by customer tags (any/all mode)
    tags: Optional[str] = None,
    tags_mode: Optional[str] = "any",
    # Sort options
    sort_by: str = "created_at",
    sort_order: str = "desc",
    limit: int = 100,
    skip: int = 0,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    and_conditions = []
    
    if search:
        and_conditions.append({
            "$or": [
                {"name": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}}
            ]
        })
    
    if tier:
        query["tier"] = tier
    
    if customer_type:
        query["customer_type"] = customer_type
    
    if has_allergies:
        query["allergies"] = {"$exists": True, "$ne": []}
    
    if last_visit_days:
        cutoff_date = (datetime.now(timezone.utc) - timedelta(days=last_visit_days)).isoformat()
        and_conditions.append({
            "$or": [
                {"last_visit": {"$lt": cutoff_date}},
                {"last_visit": None}
            ]
        })
    
    if favorite:
        query["favorites"] = {"$in": [favorite]}
    
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    
    # New filter implementations
    if whatsapp_opt_in and whatsapp_opt_in != "all":
        query["whatsapp_opt_in"] = whatsapp_opt_in == "true"
    
    if vip_flag and vip_flag != "all":
        query["vip_flag"] = vip_flag == "true"
    
    if diet_preference and diet_preference != "all":
        query["diet_preference"] = diet_preference
    
    if lead_source and lead_source != "all":
        query["lead_source"] = lead_source
    
    if preferred_time_slot and preferred_time_slot != "all":
        query["preferred_time_slot"] = preferred_time_slot
    
    if preferred_dining_type and preferred_dining_type != "all":
        query["preferred_dining_type"] = preferred_dining_type
    
    if blacklist_flag and blacklist_flag != "all":
        query["blacklist_flag"] = blacklist_flag == "true"
    
    if complaint_flag and complaint_flag != "all":
        query["complaint_flag"] = complaint_flag == "true"
    
    # Birthday this month filter
    if has_birthday_this_month:
        current_month = datetime.now(timezone.utc).month
        month_str = f"-{current_month:02d}-"
        and_conditions.append({
            "dob": {"$regex": month_str}
        })
    
    # Anniversary this month filter
    if has_anniversary_this_month:
        current_month = datetime.now(timezone.utc).month
        month_str = f"-{current_month:02d}-"
        and_conditions.append({
            "anniversary": {"$regex": month_str}
        })
    
    # Total visits filter
    if total_visits and total_visits != "all":
        if total_visits == "0":
            query["total_visits"] = 0
        elif total_visits == "1-5":
            query["total_visits"] = {"$gte": 1, "$lte": 5}
        elif total_visits == "6-10":
            query["total_visits"] = {"$gte": 6, "$lte": 10}
        elif total_visits == "10+":
            query["total_visits"] = {"$gt": 10}
    
    # Phase 3 filters
    if gender and gender != "all":
        query["gender"] = gender
    
    if total_spent and total_spent != "all":
        if total_spent == "0-500":
            query["total_spent"] = {"$gte": 0, "$lte": 500}
        elif total_spent == "500-2000":
            query["total_spent"] = {"$gt": 500, "$lte": 2000}
        elif total_spent == "2000-5000":
            query["total_spent"] = {"$gt": 2000, "$lte": 5000}
        elif total_spent == "5000-10000":
            query["total_spent"] = {"$gt": 5000, "$lte": 10000}
        elif total_spent == "10000+":
            query["total_spent"] = {"$gt": 10000}
    
    if is_blocked and is_blocked != "all":
        query["is_blocked"] = is_blocked == "true"
    
    # Quick filter: Inactive days
    if inactive_days:
        cutoff_date = (datetime.now(timezone.utc) - timedelta(days=inactive_days)).isoformat()
        and_conditions.append({
            "$or": [
                {"last_visit": {"$lt": cutoff_date}},
                {"last_visit": None}
            ]
        })
    
    # Quick filter: Most loyal (avg visits > 5 per month since registration)
    if most_loyal:
        # Use aggregation to calculate avg visits per month
        # For now, filter customers with high visit frequency
        # Customers who registered and have visits > 5 * months_since_registration
        and_conditions.append({
            "$expr": {
                "$gte": [
                    {"$divide": [
                        "$total_visits",
                        {"$max": [
                            {"$divide": [
                                {"$subtract": [{"$toDate": datetime.now(timezone.utc).isoformat()}, {"$toDate": "$created_at"}]},
                                2592000000  # milliseconds in 30 days
                            ]},
                            1
                        ]}
                    ]},
                    5
                ]
            }
        })
    
    # Feedback filter - check if customer has given feedback
    if has_feedback and has_feedback != "all":
        if has_feedback == "true":
            query["feedback_count"] = {"$gt": 0}
        else:
            query["$or"] = [{"feedback_count": {"$exists": False}}, {"feedback_count": 0}]

    # CR-043-A: tag filter — $in (any) or $all — reuses existing customers.tags array
    if tags:
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]
        if tag_list:
            if tags_mode == "all":
                query["tags"] = {"$all": tag_list}
            else:
                query["tags"] = {"$in": tag_list}

    if and_conditions:
        query["$and"] = and_conditions
    
    sort_direction = -1 if sort_order == "desc" else 1
    allowed_sort_fields = [
        "created_at", "last_visit", "total_spent", "total_points", "total_visits", 
        "name", "avg_visits_per_month", "points_balance", "wallet_balance", "tier"
    ]
    sort_field = sort_by if sort_by in allowed_sort_fields else "created_at"
    
    # For avg_visits_per_month, use total_visits as proxy (higher visits = more loyal)
    if sort_field == "avg_visits_per_month":
        sort_field = "total_visits"
    
    customers = await db.customers.find(query, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(limit).to_list(limit)
    return [Customer(**c) for c in customers]

@router.get("/segments/stats")
async def get_customer_segments(user: dict = Depends(get_current_user)):
    """Get customer segment statistics for campaign targeting"""
    user_id = user["id"]
    
    total = await db.customers.count_documents({"user_id": user_id})
    
    tier_stats = {}
    for tier in ["Bronze", "Silver", "Gold", "Platinum"]:
        tier_stats[tier.lower()] = await db.customers.count_documents({"user_id": user_id, "tier": tier})
    
    normal_count = await db.customers.count_documents({"user_id": user_id, "customer_type": "normal"})
    corporate_count = await db.customers.count_documents({"user_id": user_id, "customer_type": "corporate"})
    
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    inactive_30d = await db.customers.count_documents({
        "user_id": user_id,
        "$or": [
            {"last_visit": {"$lt": thirty_days_ago}},
            {"last_visit": None}
        ]
    })
    
    sixty_days_ago = (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()
    inactive_60d = await db.customers.count_documents({
        "user_id": user_id,
        "$or": [
            {"last_visit": {"$lt": sixty_days_ago}},
            {"last_visit": None}
        ]
    })
    
    with_allergies = await db.customers.count_documents({
        "user_id": user_id,
        "allergies": {"$exists": True, "$ne": []}
    })
    
    cities_pipeline = [
        {"$match": {"user_id": user_id, "city": {"$exists": True, "$nin": [None, ""]}}},
        {"$group": {"_id": "$city", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    cities = await db.customers.aggregate(cities_pipeline).to_list(10)
    
    favorites_pipeline = [
        {"$match": {"user_id": user_id, "favorites": {"$exists": True, "$ne": []}}},
        {"$unwind": "$favorites"},
        {"$group": {"_id": "$favorites", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_favorites = await db.customers.aggregate(favorites_pipeline).to_list(10)
    
    return {
        "total": total,
        "by_tier": tier_stats,
        "by_type": {"normal": normal_count, "corporate": corporate_count},
        "inactive_30_days": inactive_30d,
        "inactive_60_days": inactive_60d,
        "with_allergies": with_allergies,
        "top_cities": [{"city": c["_id"], "count": c["count"]} for c in cities],
        "top_favorites": [{"item": f["_id"], "count": f["count"]} for f in top_favorites]
    }

@router.get("/tags")
async def list_available_tags(
    with_counts: bool = False,             # CR-043-A: return usage counts
    user: dict = Depends(get_current_user),
):
    """CR-034 + CR-043-A: Return the tenant's tag catalog.

    Default (backward-compat): {"tags": ["VIP", "Regular", ...]} sorted alphabetically.
    with_counts=true: {"tags": [{"tag": "VIP", "count": 156}, ...]} sorted by count desc.
    """
    tenant_id = user["id"]
    user_doc = await db.users.find_one({"id": tenant_id}, {"available_tags": 1, "_id": 0})
    catalog = user_doc.get("available_tags", []) if user_doc else []

    if not with_counts:
        return {"tags": sorted(catalog)}

    # CR-043-A: aggregate customer counts per tag
    pipeline = [
        {"$match": {"user_id": tenant_id, "tags": {"$exists": True, "$ne": []}}},
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    counts = await db.customers.aggregate(pipeline).to_list(length=1000)
    counts_map = {c["_id"]: c["count"] for c in counts}

    merged = [{"tag": t, "count": counts_map.get(t, 0)} for t in catalog]
    # Include tags used on customers but not in catalog (safety net)
    for tag, count in counts_map.items():
        if tag not in catalog:
            merged.append({"tag": tag, "count": count})
    merged.sort(key=lambda x: x["count"], reverse=True)
    return {"tags": merged}


# ── CR-035: Export ────────────────────────────────────────────────────────────

@router.get("/export")
async def export_customers(
    format: str = "csv",
    user: dict = Depends(get_current_user)
):
    """CR-035: Export ALL customers as CSV or Excel. 22 fields including loyalty + tags."""
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")

    cursor = db.customers.find({"user_id": user["id"]}, {"_id": 0})
    customers = await cursor.to_list(length=None)

    def get_val(c: dict, key: str) -> str:
        v = c.get(key)
        if v is None:
            return ""
        if key == "tags" and isinstance(v, list):
            return ", ".join(v)
        if isinstance(v, bool):
            return "Yes" if v else "No"
        return str(v)

    headers = [h for h, _ in EXPORT_FIELDS]
    keys    = [k for _, k in EXPORT_FIELDS]
    timestamp = datetime.now(timezone.utc).strftime("%Y_%m_%d")

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for c in customers:
            writer.writerow([get_val(c, k) for k in keys])
        output.seek(0)
        filename = f"customers_export_{timestamp}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    header_fill = PatternFill(start_color="F26B33", end_color="F26B33", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
    for c in customers:
        ws.append([get_val(c, k) for k in keys])
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"customers_export_{timestamp}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/sample-import-template")
async def download_import_template(
    format: str = "csv",
    user: dict = Depends(get_current_user)
):
    """CR-035: Download a sample import template with headers + 2 example rows."""
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")

    IMPORT_HEADERS = ["name", "phone", "email", "dob", "city", "address", "tags"]
    SAMPLE_ROWS = [
        ["Priya Sharma",  "9876543210", "priya@example.com", "1990-05-15", "Mumbai", "123 Main St", "VIP, Regular"],
        ["Rahul Verma",   "9123456789", "rahul@example.com", "",           "Delhi",  "",             ""],
    ]
    timestamp = datetime.now(timezone.utc).strftime("%Y_%m_%d")

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(IMPORT_HEADERS)
        writer.writerows(SAMPLE_ROWS)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="import_template_{timestamp}.csv"'}
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"
    header_fill = PatternFill(start_color="F26B33", end_color="F26B33", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(IMPORT_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in SAMPLE_ROWS:
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="import_template_{timestamp}.xlsx"'}
    )


# ── CR-035: Import ────────────────────────────────────────────────────────────

@router.post("/import-preview", response_model=ImportPreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """CR-035 Step 2: Parse file and return preview — NO DB writes."""
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB.")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".csv") or fname.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported.")
    fmt = "xlsx" if fname.endswith(".xlsx") else "csv"

    rows = _parse_import_file(content, file.filename)
    if len(rows) > 5000:
        raise HTTPException(status_code=400, detail=f"File has {len(rows)} rows. Maximum allowed is 5,000.")

    existing_docs = await db.customers.find(
        {"user_id": user["id"]}, {"phone": 1, "_id": 0}
    ).to_list(length=None)
    existing_phones = {doc["phone"] for doc in existing_docs if doc.get("phone")}

    classified   = [_validate_and_classify_row(r, existing_phones) for r in rows]
    new_count    = sum(1 for r in classified if r["status"] == "new")
    update_count = sum(1 for r in classified if r["status"] == "update")
    error_count  = sum(1 for r in classified if r["status"] == "error")

    preview_rows = []
    for r in classified[:5]:
        preview_rows.append(ImportPreviewRow(
            row    = r["row"],
            name   = r.get("name"),
            phone  = r.get("phone"),
            email  = r.get("email"),
            tags   = ", ".join(r["tags"]) if isinstance(r.get("tags"), list) else r.get("tags"),
            status = r["status"],
            reason = r.get("reason"),
        ))

    all_errors = [
        ImportRowError(row=r["row"], reason=r["reason"])
        for r in classified if r["status"] == "error"
    ]

    return ImportPreviewResponse(
        filename     = file.filename,
        format       = fmt,
        total_rows   = len(rows),
        new_count    = new_count,
        update_count = update_count,
        error_count  = error_count,
        preview_rows = preview_rows,
        all_errors   = all_errors,
    )


@router.post("/import")
async def import_customers(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """CR-035 Step 3: Execute import — upsert customers, update tag catalog, log run."""
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum is 10MB.")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".csv") or fname.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx supported.")
    fmt = "xlsx" if fname.endswith(".xlsx") else "csv"

    rows = _parse_import_file(content, file.filename)
    if len(rows) > 5000:
        raise HTTPException(status_code=400, detail=f"Max 5,000 rows allowed, got {len(rows)}.")

    existing_docs = await db.customers.find(
        {"user_id": user["id"]}, {"phone": 1, "id": 1, "tags": 1, "_id": 0}
    ).to_list(length=None)
    phone_to_doc = {doc["phone"]: doc for doc in existing_docs if doc.get("phone")}

    imported_count = 0
    updated_count  = 0
    failed_count   = 0
    errors         = []
    new_tags_seen  = set()

    for raw_row in rows:
        result = _validate_and_classify_row(raw_row, set(phone_to_doc.keys()))

        if result["status"] == "error":
            failed_count += 1
            errors.append(ImportRowError(row=result["row"], reason=result["reason"]))
            continue

        payload = {"name": result["name"], "phone": result["phone"]}
        for field in ("email", "dob", "city", "address"):
            if result.get(field):
                payload[field] = result[field]

        incoming_tags = result.get("tags", [])
        for t in incoming_tags:
            new_tags_seen.add(t)

        now = datetime.now(timezone.utc).isoformat()

        if result["status"] == "update":
            existing_doc  = phone_to_doc[result["phone"]]
            existing_tags = existing_doc.get("tags", [])
            merged_tags   = list(set(existing_tags + incoming_tags))
            update_payload = {**payload, "tags": merged_tags, "updated_at": now}
            update_payload = {k: v for k, v in update_payload.items() if v is not None and v != ""}
            await db.customers.update_one(
                {"user_id": user["id"], "phone": result["phone"]},
                {"$set": update_payload}
            )
            updated_count += 1
        else:
            new_doc = {
                "id":              str(uuid.uuid4()),
                "user_id":         user["id"],
                **payload,
                "tags":            incoming_tags,
                "tier":            "Bronze",
                "total_points":    0,
                "wallet_balance":  0.0,
                "total_visits":    0,
                "total_spent":     0.0,
                "whatsapp_opt_in": False,
                "created_at":      now,
                "updated_at":      now,
            }
            await db.customers.insert_one(new_doc)
            imported_count += 1

    if new_tags_seen:
        await db.users.update_one(
            {"id": user["id"]},
            {"$addToSet": {"available_tags": {"$each": list(new_tags_seen)}}}
        )

    log = ImportLog(
        user_id    = user["id"],
        filename   = file.filename,
        format     = fmt,
        total_rows = len(rows),
        imported   = imported_count,
        updated    = updated_count,
        failed     = failed_count,
        errors     = errors[:50],
    )
    await db.import_logs.insert_one(log.dict())

    return {
        "id":         log.id,
        "filename":   log.filename,
        "total_rows": log.total_rows,
        "imported":   log.imported,
        "updated":    log.updated,
        "failed":     log.failed,
        "errors":     [e.dict() for e in errors[:50]],
        "created_at": log.created_at.isoformat(),
    }


@router.get("/import-history")
async def get_import_history(user: dict = Depends(get_current_user)):
    """CR-035: Return last 10 import logs for this tenant, newest first."""
    logs = await db.import_logs.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    return logs


@router.post("/{customer_id}/tags")
async def add_tags_to_customer(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    """CR-034: Add one or more tags to a customer. Idempotent. Updates tenant catalog."""
    import re
    new_tags = data.get("tags", [])
    if not new_tags or not isinstance(new_tags, list):
        raise HTTPException(status_code=400, detail="tags must be a non-empty list of strings")
    for t in new_tags:
        if not isinstance(t, str) or not t.strip():
            raise HTTPException(status_code=400, detail=f"Invalid tag: {t!r}")
        if len(t) > 30:
            raise HTTPException(status_code=400, detail=f"Tag '{t}' exceeds 30 characters")
        if not re.match(r'^[\w\s\-]+$', t):
            raise HTTPException(status_code=400, detail=f"Tag '{t}' contains invalid characters")
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    await db.customers.update_one(
        {"id": customer_id, "user_id": user["id"]},
        {"$addToSet": {"tags": {"$each": new_tags}}}
    )
    await db.users.update_one(
        {"id": user["id"]},
        {"$addToSet": {"available_tags": {"$each": new_tags}}}
    )
    updated = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"tags": 1, "_id": 0})
    return {"customer_id": customer_id, "tags": updated.get("tags", [])}


@router.delete("/{customer_id}/tags/{tag}")
async def remove_tag_from_customer(customer_id: str, tag: str, user: dict = Depends(get_current_user)):
    """CR-034: Remove one tag from a customer. Catalog entry kept per Q3 decision."""
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    await db.customers.update_one(
        {"id": customer_id, "user_id": user["id"]},
        {"$pull": {"tags": tag}}
    )
    updated = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"tags": 1, "_id": 0})
    return {"customer_id": customer_id, "tags": updated.get("tags", [])}


@router.post("/bulk-tag")
async def bulk_tag_customers(data: dict, user: dict = Depends(get_current_user)):
    """CR-034: Apply a tag to multiple customers in one call."""
    customer_ids = data.get("customer_ids", [])
    tag = (data.get("tag") or "").strip()
    if not customer_ids or not isinstance(customer_ids, list):
        raise HTTPException(status_code=400, detail="customer_ids must be a non-empty list")
    if not tag:
        raise HTTPException(status_code=400, detail="tag must be a non-empty string")
    if len(tag) > 30:
        raise HTTPException(status_code=400, detail="Tag exceeds 30 characters")
    result = await db.customers.update_many(
        {"id": {"$in": customer_ids}, "user_id": user["id"]},
        {"$addToSet": {"tags": tag}}
    )
    await db.users.update_one(
        {"id": user["id"]},
        {"$addToSet": {"available_tags": tag}}
    )
    return {"matched": result.matched_count, "modified": result.modified_count, "tag": tag}


@router.post("/bulk-untag")
async def bulk_untag_customers(data: dict, user: dict = Depends(get_current_user)):
    """CR-034: Remove a tag from multiple customers in one call."""
    customer_ids = data.get("customer_ids", [])
    tag = (data.get("tag") or "").strip()
    if not customer_ids or not isinstance(customer_ids, list):
        raise HTTPException(status_code=400, detail="customer_ids must be a non-empty list")
    if not tag:
        raise HTTPException(status_code=400, detail="tag must be a non-empty string")
    result = await db.customers.update_many(
        {"id": {"$in": customer_ids}, "user_id": user["id"]},
        {"$pull": {"tags": tag}}
    )
    return {"matched": result.matched_count, "modified": result.modified_count, "tag": tag}


@router.get("/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return Customer(**customer)

@router.put("/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, update_data: CustomerUpdate, request: Request, user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if "phone" in update_dict and update_dict["phone"] != customer.get("phone"):
        existing = await db.customers.find_one({
            "user_id": user["id"], 
            "phone": update_dict["phone"],
            "id": {"$ne": customer_id}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Another customer with this phone already exists")
    
    if update_dict:
        update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.customers.update_one({"id": customer_id}, {"$set": update_dict})
    
    # Sync to MyGenie if user has token (CR-008: header first, DB fallback)
    mygenie_token = request.headers.get("X-MyGenie-Token")
    if not mygenie_token:
        user_record = await db.users.find_one({"id": user["id"]})
        mygenie_token = user_record.get("mygenie_token") if user_record else None
    
    if mygenie_token:
        # Get updated customer data
        updated_customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        
        mygenie_api_url = os.environ['MYGENIE_API_URL']
        
        # Split name into first and last name
        name_parts = (updated_customer.get("name") or "").split(" ", 1)
        f_name = name_parts[0] if name_parts else ""
        l_name = name_parts[1] if len(name_parts) > 1 else ""
        
        mygenie_payload = {
            "phone": updated_customer.get("phone") or "",
            "f_name": f_name,
            "l_name": l_name,
            "email": updated_customer.get("email") or "",
            "gst_number": updated_customer.get("gst_number") or "",
            "gst_name": updated_customer.get("gst_name") or "",
            "date_of_birth": updated_customer.get("dob") or "",
            "date_of_anniversary": updated_customer.get("anniversary") or "",
            "membership_id": ""
        }
        
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.post(
                    f"{mygenie_api_url}/api/v1/vendoremployee/pos/user-check-create",
                    headers={
                        "Authorization": f"Bearer {mygenie_token}",
                        "Content-Type": "application/json; charset=UTF-8",
                        "X-localization": "en"
                    },
                    json=mygenie_payload,
                    timeout=15.0
                )
                
                if resp.status_code == 200:
                    mygenie_resp = resp.json()
                    pos_customer_id = mygenie_resp.get("user_id")
                    # Update pos_customer_id if not already set
                    if not customer.get("pos_customer_id"):
                        await db.customers.update_one(
                            {"id": customer_id},
                            {"$set": {"pos_customer_id": pos_customer_id, "mygenie_synced": True}}
                        )
                    print(f"✅ Customer updated in MyGenie: {pos_customer_id}")
                else:
                    print(f"⚠️ MyGenie update failed: {resp.status_code} - {resp.text}")
        except Exception as e:
            print(f"⚠️ MyGenie update error (non-critical): {str(e)}")
    
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return Customer(**updated)

@router.delete("/{customer_id}")
async def delete_customer(customer_id: str, user: dict = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await db.points_transactions.delete_many({"customer_id": customer_id})
    return {"message": "Customer deleted"}


# QR Code endpoints
qr_router = APIRouter(prefix="/qr", tags=["QR Code"])

@qr_router.get("/generate")
async def generate_customer_qr(user: dict = Depends(get_current_user)):
    """Generate QR code for customer registration"""
    frontend_url = os.environ['FRONTEND_URL']
    registration_url = f"{frontend_url}/register-customer/{user['id']}"
    
    qr_base64 = generate_qr_code(registration_url)
    
    return {
        "qr_code": f"data:image/png;base64,{qr_base64}",
        "registration_url": registration_url,
        "restaurant_name": user["restaurant_name"]
    }

@qr_router.post("/register/{restaurant_id}")
async def register_via_qr(restaurant_id: str, customer_data: CustomerCreate):
    """Register customer via QR code (no auth required)"""
    user = await db.users.find_one({"id": restaurant_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    existing = await db.customers.find_one({"user_id": restaurant_id, "phone": customer_data.phone})
    if existing:
        raise HTTPException(status_code=400, detail="Customer already registered")
    
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Check for first visit bonus
    settings = await db.loyalty_settings.find_one({"user_id": restaurant_id}, {"_id": 0})
    first_visit_bonus = 0
    if settings and settings.get("first_visit_bonus_enabled", False):
        first_visit_bonus = settings.get("first_visit_bonus_points", 50)
    
    customer_doc = {
        "id": customer_id,
        "user_id": restaurant_id,
        "created_at": now,
        "updated_at": now,
        
        # Basic Information
        "name": customer_data.name,
        "phone": customer_data.phone,
        "country_code": customer_data.country_code,
        "email": customer_data.email,
        "gender": customer_data.gender,
        "dob": customer_data.dob,
        "anniversary": customer_data.anniversary,
        "preferred_language": customer_data.preferred_language,
        "customer_type": customer_data.customer_type,
        "segment_tags": customer_data.segment_tags or [],
        
        # Contact & Marketing Permissions
        "whatsapp_opt_in": customer_data.whatsapp_opt_in,
        "whatsapp_opt_in_date": customer_data.whatsapp_opt_in_date,
        "promo_whatsapp_allowed": customer_data.promo_whatsapp_allowed,
        "promo_sms_allowed": customer_data.promo_sms_allowed,
        "email_marketing_allowed": customer_data.email_marketing_allowed,
        "call_allowed": customer_data.call_allowed,
        "is_blocked": customer_data.is_blocked,
        
        # Loyalty Information
        "total_points": first_visit_bonus,
        "wallet_balance": 0.0,
        "tier": "Bronze",
        "referral_code": customer_data.referral_code,
        "referred_by": customer_data.referred_by,
        "membership_id": customer_data.membership_id,
        "membership_expiry": customer_data.membership_expiry,
        
        # Spending & Visit Behavior
        "total_visits": 0,
        "total_spent": 0.0,
        "avg_order_value": 0.0,
        "last_visit": None,
        "first_visit_date": now,
        "favorite_category": customer_data.favorite_category,
        "preferred_payment_mode": customer_data.preferred_payment_mode,
        
        # Customer Source & Journey
        "lead_source": customer_data.lead_source,
        "campaign_source": customer_data.campaign_source,
        "last_interaction_date": now,
        "assigned_salesperson": customer_data.assigned_salesperson,
        
        # WhatsApp CRM Tracking
        "last_whatsapp_sent": None,
        "last_whatsapp_response": None,
        "last_campaign_clicked": None,
        "last_coupon_used": None,
        "automation_status_tag": None,
        
        # Corporate Information
        "gst_name": customer_data.gst_name,
        "gst_number": customer_data.gst_number,
        "billing_address": customer_data.billing_address,
        "credit_limit": customer_data.credit_limit,
        "payment_terms": customer_data.payment_terms,
        
        # Address
        "address": customer_data.address,
        "address_line_2": customer_data.address_line_2,
        "city": customer_data.city,
        "state": customer_data.state,
        "pincode": customer_data.pincode,
        "country": customer_data.country,
        "delivery_instructions": customer_data.delivery_instructions,
        "map_location": customer_data.map_location,
        
        # Preferences
        "allergies": customer_data.allergies or [],
        "favorites": customer_data.favorites or [],
        
        # Dining Preferences
        "preferred_dining_type": customer_data.preferred_dining_type,
        "preferred_time_slot": customer_data.preferred_time_slot,
        "favorite_table": customer_data.favorite_table,
        "avg_party_size": customer_data.avg_party_size,
        "diet_preference": customer_data.diet_preference,
        "spice_level": customer_data.spice_level,
        "cuisine_preference": customer_data.cuisine_preference,
        
        # Special Occasions
        "kids_birthday": customer_data.kids_birthday or [],
        "spouse_name": customer_data.spouse_name,
        "festival_preference": customer_data.festival_preference or [],
        "special_dates": customer_data.special_dates or [],
        
        # Feedback & Flags
        "last_rating": customer_data.last_rating,
        "nps_score": customer_data.nps_score,
        "complaint_flag": customer_data.complaint_flag,
        "vip_flag": customer_data.vip_flag,
        "blacklist_flag": customer_data.blacklist_flag,
        
        # AI/Advanced
        "predicted_next_visit": customer_data.predicted_next_visit,
        "churn_risk_score": customer_data.churn_risk_score,
        "recommended_offer_type": customer_data.recommended_offer_type,
        "price_sensitivity_score": customer_data.price_sensitivity_score,
        
        # Custom Fields
        "custom_field_1": customer_data.custom_field_1,
        "custom_field_2": customer_data.custom_field_2,
        "custom_field_3": customer_data.custom_field_3,
        
        # Notes
        "notes": customer_data.notes,
        
        # Bonus
        "first_visit_bonus_awarded": first_visit_bonus > 0
    }
    
    await db.customers.insert_one(customer_doc)
    
    # Record first visit bonus transaction if awarded
    if first_visit_bonus > 0:
        tx_doc = {
            "id": str(uuid.uuid4()),
            "user_id": restaurant_id,
            "customer_id": customer_id,
            "points": first_visit_bonus,
            "transaction_type": "bonus",
            "description": "First visit bonus - Welcome reward",
            "bill_amount": None,
            "balance_after": first_visit_bonus,
            "created_at": now
        }
        await db.points_transactions.insert_one(tx_doc)
    
    return {
        "message": "Registration successful",
        "customer_id": customer_id,
        "first_visit_bonus_awarded": first_visit_bonus
    }


# Segments router
segments_router = APIRouter(prefix="/segments", tags=["Segments"])

async def count_customers_by_filters(user_id: str, filters: dict) -> int:
    query = await build_customer_query(user_id, filters)
    return await db.customers.count_documents(query)

@segments_router.post("", response_model=Segment)
async def create_segment(segment_data: SegmentCreate, user: dict = Depends(get_current_user)):
    segment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Use frontend-provided count if available, otherwise calculate
    if segment_data.customer_count is not None:
        customer_count = segment_data.customer_count
    else:
        customer_count = await count_customers_by_filters(user["id"], segment_data.filters)
    
    segment_doc = {
        "id": segment_id,
        "user_id": user["id"],
        "name": segment_data.name,
        "filters": segment_data.filters,
        "customer_count": customer_count,
        "created_at": now,
        "updated_at": now
    }
    
    await db.segments.insert_one(segment_doc)
    return Segment(**segment_doc)

@segments_router.get("", response_model=List[Segment])
async def list_segments(user: dict = Depends(get_current_user)):
    """CR-024 Phase 4 P4.2: Return cached counts (no per-list recount).
    Use POST /segments/{id}/refresh-count for an explicit refresh.
    Daily cron also refreshes counts in the background.
    """
    segments = await db.segments.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    return [Segment(**s) for s in segments]

@segments_router.post("/{segment_id}/refresh-count")
async def refresh_segment_count(segment_id: str, user: dict = Depends(get_current_user)):
    """CR-024 Phase 4 P4.2: Explicit recount + persist last_counted_at."""
    segment = await db.segments.find_one({"id": segment_id, "user_id": user["id"]})
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    count = await count_customers_by_filters(user["id"], segment["filters"])
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.segments.update_one(
        {"id": segment_id},
        {"$set": {"customer_count": count, "last_counted_at": now_iso}},
    )
    return {"segment_id": segment_id, "customer_count": count, "last_counted_at": now_iso}

@segments_router.post("/preview-count")
async def preview_segment_count(data: dict, user: dict = Depends(get_current_user)):
    """Preview customer count for a set of filters before creating a segment"""
    filters = data.get("filters", {})
    count = await count_customers_by_filters(user["id"], filters)
    return {"count": count}

@segments_router.get("/{segment_id}", response_model=Segment)
async def get_segment(segment_id: str, user: dict = Depends(get_current_user)):
    # CR-024 Phase 4 P4.3: synthetic "all-customers" support
    if segment_id == "all-customers":
        count = await db.customers.count_documents({"user_id": user["id"]})
        now_iso = datetime.now(timezone.utc).isoformat()
        return Segment(
            id="all-customers", user_id=user["id"], name="All Customers",
            filters={}, customer_count=count, last_counted_at=now_iso,
            created_at=now_iso, updated_at=now_iso,
        )

    segment = await db.segments.find_one({"id": segment_id, "user_id": user["id"]}, {"_id": 0})
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    
    count = await count_customers_by_filters(user["id"], segment["filters"])
    segment["customer_count"] = count
    
    return Segment(**segment)

@segments_router.get("/{segment_id}/customers", response_model=List[Customer])
async def get_segment_customers(segment_id: str, user: dict = Depends(get_current_user)):
    # CR-024 Phase 4 P4.3: synthetic "all-customers" support
    if segment_id == "all-customers":
        customers = await db.customers.find({"user_id": user["id"]}, {"_id": 0}).to_list(1000)
        return [Customer(**c) for c in customers]

    segment = await db.segments.find_one({"id": segment_id, "user_id": user["id"]}, {"_id": 0})
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    
    query = await build_customer_query(user["id"], segment["filters"])
    customers = await db.customers.find(query, {"_id": 0}).to_list(1000)
    
    return [Customer(**c) for c in customers]

@segments_router.put("/{segment_id}", response_model=Segment)
async def update_segment(segment_id: str, update_data: SegmentUpdate, user: dict = Depends(get_current_user)):
    segment = await db.segments.find_one({"id": segment_id, "user_id": user["id"]})
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict:
        update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        if "filters" in update_dict:
            count = await count_customers_by_filters(user["id"], update_dict["filters"])
            update_dict["customer_count"] = count
            update_dict["last_counted_at"] = update_dict["updated_at"]
        
        await db.segments.update_one({"id": segment_id}, {"$set": update_dict})
    
    updated_segment = await db.segments.find_one({"id": segment_id}, {"_id": 0})
    return Segment(**updated_segment)

@segments_router.delete("/{segment_id}")
async def delete_segment(segment_id: str, user: dict = Depends(get_current_user)):
    result = await db.segments.delete_one({"id": segment_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Segment not found")
    # Also delete any WhatsApp config for this segment
    await db.segment_whatsapp_config.delete_one({"segment_id": segment_id, "user_id": user["id"]})
    return {"message": "Segment deleted"}

# WhatsApp Configuration for Segments
@segments_router.get("/{segment_id}/whatsapp-config")
async def get_segment_whatsapp_config(segment_id: str, user: dict = Depends(get_current_user)):
    """Get WhatsApp automation config for a segment"""
    config = await db.segment_whatsapp_config.find_one(
        {"segment_id": segment_id, "user_id": user["id"]},
        {"_id": 0}
    )
    if not config:
        return {"configured": False}
    return {"configured": True, "config": config}

@segments_router.post("/{segment_id}/whatsapp-config")
async def save_segment_whatsapp_config(segment_id: str, config: dict, user: dict = Depends(get_current_user)):
    """Save WhatsApp automation config for a segment"""
    from datetime import datetime, timezone
    
    # Verify segment exists
    segment = await db.segments.find_one({"id": segment_id, "user_id": user["id"]})
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    
    now = datetime.now(timezone.utc).isoformat()
    config_doc = {
        "segment_id": segment_id,
        "user_id": user["id"],
        "template_id": config.get("template_id"),
        "template_name": config.get("template_name"),
        "variable_mappings": config.get("variable_mappings", {}),
        "variable_modes": config.get("variable_modes", {}),
        "schedule_type": config.get("schedule_type", "now"),  # now, scheduled, recurring
        "scheduled_date": config.get("scheduled_date"),
        "scheduled_time": config.get("scheduled_time", "10:00"),
        "recurring_frequency": config.get("recurring_frequency"),  # daily, weekly, monthly
        "recurring_days": config.get("recurring_days", []),
        "recurring_day_of_month": config.get("recurring_day_of_month"),
        "recurring_end_option": config.get("recurring_end_option", "never"),
        "recurring_end_date": config.get("recurring_end_date"),
        "recurring_occurrences": config.get("recurring_occurrences"),
        "is_active": True,
        "created_at": now,
        "updated_at": now
    }
    
    # Upsert the config
    await db.segment_whatsapp_config.update_one(
        {"segment_id": segment_id, "user_id": user["id"]},
        {"$set": config_doc},
        upsert=True
    )
    
    return {"message": "WhatsApp config saved", "config": config_doc}

@segments_router.delete("/{segment_id}/whatsapp-config")
async def delete_segment_whatsapp_config(segment_id: str, user: dict = Depends(get_current_user)):
    """Remove WhatsApp automation config for a segment"""
    result = await db.segment_whatsapp_config.delete_one(
        {"segment_id": segment_id, "user_id": user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="WhatsApp config not found")
    return {"message": "WhatsApp config removed"}

@segments_router.patch("/{segment_id}/whatsapp-config/toggle")
async def toggle_segment_whatsapp_config(segment_id: str, user: dict = Depends(get_current_user)):
    """Pause or resume WhatsApp automation for a segment"""
    from datetime import datetime, timezone
    
    config = await db.segment_whatsapp_config.find_one(
        {"segment_id": segment_id, "user_id": user["id"]}
    )
    if not config:
        raise HTTPException(status_code=404, detail="WhatsApp config not found")
    
    new_status = not config.get("is_active", True)
    now = datetime.now(timezone.utc).isoformat()
    
    await db.segment_whatsapp_config.update_one(
        {"segment_id": segment_id, "user_id": user["id"]},
        {"$set": {"is_active": new_status, "updated_at": now}}
    )
    
    return {
        "message": f"WhatsApp automation {'resumed' if new_status else 'paused'}",
        "is_active": new_status
    }

@segments_router.get("/whatsapp-configs/all")
async def get_all_segment_whatsapp_configs(user: dict = Depends(get_current_user)):
    """Get all WhatsApp configs for user's segments"""
    configs = await db.segment_whatsapp_config.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).to_list(100)
    return {"configs": configs}


@router.get("/{customer_id}/loyalty-details")
async def get_customer_loyalty_details(customer_id: str, user: dict = Depends(get_current_user)):
    """Get loyalty conversion rate and coupon summary for a customer"""
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Get loyalty settings for conversion rate
    loyalty_settings = await db.loyalty_settings.find_one({"user_id": user["id"]}, {"_id": 0})
    redemption_value = 0.25  # default: 1 point = ₹0.25
    if loyalty_settings:
        redemption_value = loyalty_settings.get("redemption_value", 0.25)

    # Get active coupons for the restaurant
    now = datetime.now(timezone.utc).isoformat()
    active_coupons = await db.coupons.find(
        {"user_id": user["id"], "is_active": True},
        {"_id": 0, "id": 1, "code": 1, "description": 1, "discount_type": 1, "discount_value": 1, "max_discount": 1, "valid_until": 1, "end_date": 1, "usage_limit": 1, "used_count": 1, "total_used": 1}
    ).to_list(50)

    # Calculate monetary values
    total_points = customer.get("total_points", 0)
    total_earned = customer.get("total_points_earned", 0)
    total_redeemed = customer.get("total_points_redeemed", 0)

    return {
        "redemption_value": redemption_value,
        "points_money_value": round(total_points * redemption_value, 2),
        "earned_money_value": round(total_earned * redemption_value, 2),
        "redeemed_money_value": round(total_redeemed * redemption_value, 2),
        "total_coupon_used": customer.get("total_coupon_used", 0),
        "active_coupons": active_coupons
    }


@router.get("/{customer_id}/coupon-history")
async def get_customer_coupon_history(customer_id: str, user: dict = Depends(get_current_user)):
    """CR-002B: Returns this customer's coupon usage history with discount, date, order_id."""
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"_id": 0, "id": 1})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    usages = await db.coupon_usage.find(
        {"customer_id": customer_id, "user_id": user["id"]},
        {"_id": 0, "id": 1, "coupon_code": 1, "coupon_title": 1, "discount_scope": 1,
         "coupon_discount": 1, "order_id": 1, "pos_order_id": 1, "used_at": 1,
         "created_at": 1, "offer_type": 1, "discount_type": 1, "discount_value": 1}
    ).sort("created_at", -1).limit(50).to_list(50)

    return {"customer_id": customer_id, "coupon_usages": usages, "total": len(usages)}


@router.get("/{customer_id}/insights")
async def get_customer_insights(customer_id: str, user: dict = Depends(get_current_user)):
    """P0 AI Insights - aggregation-based, no ML required"""
    customer = await db.customers.find_one({"id": customer_id, "user_id": user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    insights = {}

    # 1. Top Items (from order_items)
    top_items_pipeline = [
        {"$match": {"customer_id": customer_id, "user_id": user["id"]}},
        {"$group": {"_id": "$item_name", "count": {"$sum": "$item_qty"}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    top_items = await db.order_items.aggregate(top_items_pipeline).to_list(5)
    insights["top_items"] = [{"name": i["_id"], "count": i["count"]} for i in top_items]

    # 2. Preferred Category (from order_items)
    category_pipeline = [
        {"$match": {"customer_id": customer_id, "user_id": user["id"], "item_category": {"$ne": None}}},
        {"$group": {"_id": "$item_category", "count": {"$sum": "$item_qty"}}},
        {"$sort": {"count": -1}},
        {"$limit": 3}
    ]
    categories = await db.order_items.aggregate(category_pipeline).to_list(3)
    total_cat = sum(c["count"] for c in categories) if categories else 0
    insights["top_categories"] = [
        {"name": c["_id"], "count": c["count"], "percent": round(c["count"] / total_cat * 100) if total_cat > 0 else 0}
        for c in categories
    ]

    # 3. Order Frequency & Preferred Day/Time (from orders)
    orders = await db.orders.find(
        {"customer_id": customer_id, "user_id": user["id"]},
        {"_id": 0, "created_at": 1, "order_amount": 1}
    ).sort("created_at", 1).to_list(1000)

    if len(orders) >= 2:
        dates = []
        for o in orders:
            try:
                dt = datetime.fromisoformat(o["created_at"].replace("Z", "+00:00")) if isinstance(o["created_at"], str) else o["created_at"]
                dates.append(dt)
            except Exception:
                pass

        if len(dates) >= 2:
            dates_naive = [dt.replace(tzinfo=None) for dt in dates]
            gaps = [(dates_naive[i+1] - dates_naive[i]).days for i in range(len(dates_naive)-1) if (dates_naive[i+1] - dates_naive[i]).days > 0]
            insights["avg_frequency_days"] = round(sum(gaps) / len(gaps)) if gaps else None

            # Preferred day of week
            day_counts = {}
            for dt in dates:
                day_name = dt.strftime("%A")
                day_counts[day_name] = day_counts.get(day_name, 0) + 1
            if day_counts:
                preferred_day = max(day_counts, key=day_counts.get)
                insights["preferred_day"] = preferred_day

            # Preferred time slot
            hour_slots = {"Breakfast (8-11 AM)": 0, "Lunch (12-3 PM)": 0, "Evening (4-7 PM)": 0, "Dinner (7-11 PM)": 0, "Late Night (11 PM+)": 0}
            for dt in dates:
                h = dt.hour
                if 8 <= h < 11:
                    hour_slots["Breakfast (8-11 AM)"] += 1
                elif 12 <= h < 15:
                    hour_slots["Lunch (12-3 PM)"] += 1
                elif 16 <= h < 19:
                    hour_slots["Evening (4-7 PM)"] += 1
                elif 19 <= h < 23:
                    hour_slots["Dinner (7-11 PM)"] += 1
                else:
                    hour_slots["Late Night (11 PM+)"] += 1
            active_slots = {k: v for k, v in hour_slots.items() if v > 0}
            if active_slots:
                insights["preferred_time"] = max(active_slots, key=active_slots.get)
        else:
            insights["avg_frequency_days"] = None
    else:
        insights["avg_frequency_days"] = None

    # 4. Spending Trend (last 3 months vs previous 3 months)
    if len(orders) >= 3:
        now = datetime.now(timezone.utc)
        recent = []
        older = []
        for o in orders:
            try:
                dt = datetime.fromisoformat(o["created_at"].replace("Z", "+00:00")) if isinstance(o["created_at"], str) else o["created_at"]
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days_ago = (now - dt).days
                if days_ago <= 90:
                    recent.append(o["order_amount"])
                elif days_ago <= 180:
                    older.append(o["order_amount"])
            except Exception:
                pass

        if recent and older:
            recent_avg = sum(recent) / len(recent)
            older_avg = sum(older) / len(older)
            if older_avg > 0:
                change = round((recent_avg - older_avg) / older_avg * 100)
                insights["spending_trend"] = {"change_percent": change, "direction": "up" if change > 0 else "down"}

    # 5. Common Customizations (from item_notes)
    notes_pipeline = [
        {"$match": {"customer_id": customer_id, "user_id": user["id"], "item_notes": {"$ne": None}}},
        {"$group": {"_id": "$item_notes", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    notes = await db.order_items.aggregate(notes_pipeline).to_list(5)
    insights["common_notes"] = [{"note": n["_id"], "count": n["count"]} for n in notes]

    # 6. Avg order value
    insights["avg_order_value"] = customer.get("avg_order_value", 0)

    return insights


